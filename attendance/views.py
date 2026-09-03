from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db import transaction
from django.db.models import F
from django.core.exceptions import PermissionDenied
from functools import wraps
from .models import Shift, ShiftAssignment, AttendanceCorrection
from django.core.exceptions import ValidationError
from employees.models import Employee
from django.shortcuts import redirect

def staff_required(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if not request.user.is_staff:
            raise PermissionDenied("Only authorized staff can access this page.")
        return view_func(request, *args, **kwargs)
    return _wrapped_view

from django.db.models import Q
from django.contrib.auth.decorators import login_required
from django.conf import settings
from django.urls import reverse
from datetime import datetime, date, timedelta, time
import calendar
from django.utils import timezone
from .models import Attendance


def _month_bounds(anchor_date):
    month_start = anchor_date.replace(day=1)
    month_end = anchor_date.replace(day=calendar.monthrange(anchor_date.year, anchor_date.month)[1])
    return month_start, month_end


def _attendance_row(employee, row_date, attendance=None, is_staff=False, today=None, holiday=None, leave_info=None):
    today = today or timezone.localdate()
    correction = None
    if attendance:
        corrections = list(attendance.corrections.all())
        correction = corrections[0] if corrections else None

    if attendance and attendance.clock_in and attendance.clock_out:
        # Calculate hours if net_work_hours is missing (e.g., from imports)
        worked_hours = attendance.net_work_hours
        if worked_hours is None:
            import datetime
            t1 = datetime.datetime.combine(attendance.date, attendance.clock_in)
            t2 = datetime.datetime.combine(attendance.date, attendance.clock_out)
            worked_hours = (t2 - t1).total_seconds() / 3600.0
            
        if worked_hours is not None and 3 <= float(worked_hours) < 7:
            status_label = "Half Day"
            status_type = "half_day"
        else:
            status_label = "Completed"
            status_type = "completed"
    elif attendance and attendance.clock_in and row_date == today:
        status_label = "Working"
        status_type = "working"
    elif attendance and attendance.clock_in:
        status_label = "Missed Out"
        status_type = "missing"
    elif leave_info:
        if leave_info.get('is_half_day'):
            status_label = "Half Day Leave"
            status_type = "half_leave"
        else:
            status_label = "On Leave"
            status_type = "leave"
    elif holiday:
        status_label = "Holiday"
        status_type = "holiday"
    elif row_date.weekday() == 6:  # Sunday only
        status_label = "Weekly Off"
        status_type = "weekend"
    elif row_date > today:
        status_label = "Upcoming"
        status_type = "future"
    else:
        status_label = "Absent"
        status_type = "absent"

    return {
        "employee": employee,
        "date": row_date,
        "attendance": attendance,
        "id": attendance.id if attendance else "",
        "clock_in": attendance.clock_in if attendance else None,
        "clock_out": attendance.clock_out if attendance else None,
        "late_minutes": attendance.late_minutes if attendance else 0,
        "early_out_minutes": attendance.early_out_minutes if attendance else 0,
        "formatted_late_time": attendance.formatted_late_time if attendance else "",
        "formatted_early_out_time": attendance.formatted_early_out_time if attendance else "",
        "current_work_time": attendance.current_work_time if attendance else "--",
        "total_break_minutes": attendance.total_break_minutes if attendance else 0,
        "net_work_hours": attendance.net_work_hours if attendance else None,
        "status_label": status_label,
        "status_type": status_type,
        "correction": correction,
        "can_request_correction": bool(attendance and not is_staff),
        "admin_adjusted": bool(attendance and attendance.admin_actions.exists()),
    }


def _build_visual_attendance_register(employees, month_start, month_end, attendance_map, organization, today, go_live_date=None):
    from leaves.models import Holiday, LeaveRequest

    register_dates = [
        month_start + timedelta(days=offset)
        for offset in range((month_end - month_start).days + 1)
    ]

    holiday_qs = Holiday.objects.filter(
        organization=organization,
        date__range=(month_start, month_end),
        is_deleted=False,
    ).select_related('calendar')
    holiday_map = {holiday.date: holiday for holiday in holiday_qs}

    leave_day_map = {}
    if employees:
        leave_qs = LeaveRequest.objects.filter(
            organization=organization,
            employee__in=employees,
            status='APPROVED',
            start_date__lte=month_end,
            end_date__gte=month_start,
            is_deleted=False,
        ).select_related('employee', 'leave_type')

        for leave in leave_qs:
            leave_day = max(leave.start_date, month_start)
            leave_end = min(leave.end_date, month_end)
            is_half_day = (
                leave.session_type in ('MORNING', 'AFTERNOON', 'SHORT')
                or float(leave.total_days or 0) < 1
            )
            while leave_day <= leave_end:
                leave_day_map[(leave.employee_id, leave_day)] = {
                    'class': 'half' if is_half_day else 'leave',
                    'code': 'H' if is_half_day else 'L',
                    'label': 'Half Day' if is_half_day else 'Leave',
                    'title': f"{leave.leave_type.name} ({leave.get_session_type_display()})",
                }
                leave_day += timedelta(days=1)

    counts = {
        'present': 0,
        'leave': 0,
        'half': 0,
        'absent': 0,
        'holiday': 0,
        'weekend': 0,
    }
    rows = []

    for employee in employees:
        cells = []
        for row_date in register_dates:
            attendance = attendance_map.get((employee.id, row_date))
            leave_cell = leave_day_map.get((employee.id, row_date))
            holiday = holiday_map.get(row_date)
            is_weekend = row_date.weekday() == 6  # Sunday only
            title_date = row_date.strftime('%d %b %Y')

            # Days before go-live with NO attendance data → show as not tracked (neutral)
            # Days before go-live WITH imported attendance data → show normally
            is_before_go_live = go_live_date and row_date < go_live_date and not attendance

            if leave_cell:
                cell = {
                    'date': row_date,
                    'class': leave_cell['class'],
                    'code': leave_cell['code'],
                    'label': leave_cell['label'],
                    'title': f"{title_date}: {leave_cell['title']}",
                    'is_today': row_date == today,
                }
                counts['half' if leave_cell['class'] == 'half' else 'leave'] += 1
            elif attendance and attendance.clock_in:
                cell = {
                    'date': row_date,
                    'class': 'present',
                    'code': 'P',
                    'label': 'Present',
                    'title': f"{title_date}: In {attendance.clock_in.strftime('%I:%M %p')}"
                             f"{' / Out ' + attendance.clock_out.strftime('%I:%M %p') if attendance.clock_out else ''}",
                    'is_today': row_date == today,
                }
                counts['present'] += 1
            elif holiday:
                cell = {
                    'date': row_date,
                    'class': 'holiday',
                    'code': 'OH' if holiday.is_optional else 'HD',
                    'label': 'Optional Holiday' if holiday.is_optional else 'Holiday',
                    'title': f"{title_date}: {holiday.name}",
                    'is_today': row_date == today,
                }
                counts['holiday'] += 1
            elif is_weekend:
                cell = {
                    'date': row_date,
                    'class': 'weekend',
                    'code': 'W',
                    'label': 'Weekly Off',
                    'title': f"{title_date}: Weekly off",
                    'is_today': row_date == today,
                }
                counts['weekend'] += 1
            elif is_before_go_live:
                # Before system go-live with no data: show as not tracked (not counted as absent)
                cell = {
                    'date': row_date,
                    'class': 'future',
                    'code': '-',
                    'label': 'Not Tracked',
                    'title': f"{title_date}: Before system go-live",
                    'is_today': False,
                }
            elif row_date <= today:
                cell = {
                    'date': row_date,
                    'class': 'absent',
                    'code': 'A',
                    'label': 'Absent',
                    'title': f"{title_date}: No attendance marked",
                    'is_today': row_date == today,
                }
                counts['absent'] += 1
            else:
                cell = {
                    'date': row_date,
                    'class': 'future',
                    'code': '-',
                    'label': 'Upcoming',
                    'title': f"{title_date}: Upcoming day",
                    'is_today': False,
                }

            cells.append(cell)

        rows.append({
            'employee': employee,
            'days': cells,
        })

    return {
        'days': [{'date': item, 'is_today': item == today} for item in register_dates],
        'rows': rows,
        'counts': counts,
    }


@login_required
def attendance_view(request):
    current_employee = Employee.objects.filter(email=request.user.email, organization=request.user.organization, is_active=True, is_deleted=False).first()
    
    if not request.user.is_staff and not current_employee:
        messages.error(request, "You do not have permission to view the Attendance Dashboard.")
        return redirect('home')
        
    organization = request.user.organization
    today = timezone.localdate()
    
    # --- Date Resolution ---
    # Priority: ?date= param (single-day) > ?month= param (legacy) > today
    filter_date_str = request.GET.get('date', '')
    month_str = request.GET.get('month', '')
    view_mode = request.GET.get('view', 'list')
    if view_mode not in ('list', 'visual'):
        view_mode = 'list'

    if filter_date_str:
        try:
            selected_date = datetime.strptime(filter_date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = today
    elif month_str:
        # Legacy month param – show first day of that month
        try:
            selected_date = datetime.strptime(f"{month_str}-01", '%Y-%m-%d').date()
        except ValueError:
            selected_date = today
    else:
        # DEFAULT: today only
        selected_date = today

    display_date = selected_date
    month_start, month_end = _month_bounds(selected_date)
    current_month_start = today.replace(day=1)

    # Resolve HRMS_GO_LIVE_DATE – used only to mark pre-go-live empty cells as
    # "not tracked" instead of "absent". It no longer cuts off the display range,
    # so imported historical data for dates before the go-live date still shows.
    _go_live_date = None
    _go_live_str = getattr(settings, 'HRMS_GO_LIVE_DATE', '')
    if _go_live_str:
        try:
            _go_live_date = datetime.strptime(_go_live_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    employee_qs = Employee.objects.filter(
        organization=organization,
        is_active=True,
        is_deleted=False,
    ).select_related('department').order_by('first_name', 'last_name', 'employee_id')

    search_query = request.GET.get('search', '')
    if search_query and request.user.is_staff:
        employee_qs = employee_qs.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(employee_id__icontains=search_query)
        )
        
    if not request.user.is_staff:
        employee_qs = employee_qs.filter(pk=current_employee.pk)

    employees = list(employee_qs)

    # Load attendance for the entire month (needed for visual register)
    # but build the list view for selected_date ONLY
    attendance_map = {}
    if employees:
        attendance_qs = Attendance.objects.filter(
            employee__in=employees,
            date__range=(month_start, month_end),
        ).select_related(
            'employee',
            'employee__department',
            'shift',
        ).prefetch_related('corrections', 'admin_actions')
        attendance_map = {(att.employee_id, att.date): att for att in attendance_qs}

    # Single-day attendance rows for the selected date
    from leaves.models import Holiday, LeaveRequest
    
    # Fetch holidays for the selected date
    holiday = Holiday.objects.filter(
        organization=organization,
        date=selected_date,
        is_deleted=False
    ).first()

    # Fetch approved leaves that span the selected date
    leave_qs = LeaveRequest.objects.filter(
        employee__in=employees,
        status='APPROVED',
        start_date__lte=selected_date,
        end_date__gte=selected_date,
        is_deleted=False,
    ).select_related('leave_type')
    
    # Create a map for fast lookup: employee.id -> LeaveRequest
    leave_map = {}
    for lr in leave_qs:
        leave_map[lr.employee_id] = {
            'leave_type': lr.leave_type.name if lr.leave_type else 'Leave',
            'is_half_day': lr.session_type != 'FULL'
        }

    attendance_rows = []
    for employee in employees:
        attendance_rows.append(
            _attendance_row(
                employee,
                selected_date,
                attendance_map.get((employee.id, selected_date)),
                is_staff=request.user.is_staff,
                today=today,
                holiday=holiday,
                leave_info=leave_map.get(employee.id)
            )
        )

    visual_register = _build_visual_attendance_register(
        employees=employees,
        month_start=month_start,  # Always show the full month — go_live_date only hides untracked empty cells
        month_end=month_end,
        attendance_map=attendance_map,
        organization=organization,
        today=today,
        go_live_date=_go_live_date,
    )

    # Stats for selected date only
    total_employees = Employee.objects.filter(organization=organization, is_active=True, is_deleted=False).count()
    present_count = sum(1 for row in attendance_rows if row["clock_in"])
    late_count = sum(1 for row in attendance_rows if row["late_minutes"] > 0)
    absent_count = sum(1 for row in attendance_rows if row["status_type"] == "absent")
    
    # Attendance rate (present / total, expressed as percentage)
    attendance_rate = round((present_count / total_employees * 100), 1) if total_employees > 0 else 0

    pending_corrections_count = 0
    if request.user.is_staff:
        from .models import AttendanceCorrection
        pending_corrections_count = AttendanceCorrection.objects.filter(employee__organization=organization, status='PENDING').count()

    # Prev / next date for navigation
    prev_date = selected_date - timedelta(days=1)
    next_date = selected_date + timedelta(days=1)
    
    # Prev / next month for visual view
    prev_month_date = (month_start - timedelta(days=1)).replace(day=1)
    next_month_date = (month_end + timedelta(days=1))
    
    context = {
        'attendances': attendance_rows,
        'total_employees': total_employees,
        'present_count': present_count,
        'absent_count': absent_count,
        'late_count': late_count,
        'attendance_rate': attendance_rate,
        'display_date': display_date,
        'selected_date': selected_date,
        'prev_date': prev_date,
        'next_date': next_date,
        'month_start': month_start,
        'month_end': month_end,
        'selected_month_value': month_start.strftime('%Y-%m'),
        'prev_month_value': prev_month_date.strftime('%Y-%m'),
        'next_month_value': next_month_date.strftime('%Y-%m'),
        'today': today,
        'search_query': search_query,
        'pending_corrections_count': pending_corrections_count,
        'attendance_register_days': visual_register['days'],
        'attendance_register_rows': visual_register['rows'],
        'attendance_register_counts': visual_register['counts'],
        'view_mode': view_mode,
        'is_today': selected_date == today,
    }
    return render(request, 'attendance_dashboard.html', context)



@login_required
def attendance_visual_view(request):
    params = request.GET.copy()
    params['view'] = 'visual'
    return redirect(f"{reverse('attendance')}?{params.urlencode()}")


@login_required
@staff_required
def create_shift_view(request,pk=None):
    edit_shift = None
    
    # Handle edit query parameter
    if not pk and request.GET.get('edit'):
        pk = request.GET.get('edit')
        
    if pk:
        try:
            edit_shift = Shift.objects.get(pk=pk, organization=request.user.organization, is_deleted=False)
        except Shift.DoesNotExist:
            messages.error(request, 'Shift not found.')
            return redirect('create_shift')
            
    if request.method == 'POST':
        # Handle delete
        delete_id = request.GET.get('delete')
        if delete_id:
            try:
                shift_to_delete = Shift.objects.get(pk=delete_id, organization=request.user.organization, is_deleted=False)
                shift_to_delete.is_deleted = True
                shift_to_delete.updated_by = request.user
                shift_to_delete.save()
                messages.success(request, f'Shift "{shift_to_delete.name}" deleted successfully.')
            except Shift.DoesNotExist:
                messages.error(request, 'Shift not found.')
            return redirect('create_shift')
            
        name = request.POST.get('name', '').strip()
        start_time = request.POST.get('start_time', '').strip()
        end_time = request.POST.get('end_time', '').strip()
        grace_minutes = request.POST.get('grace_minutes', '').strip()
        minimum_half_day_hours = request.POST.get('minimum_half_day_hours', '').strip()
        minimum_full_day_hours = request.POST.get('minimum_full_day_hours', '').strip()
        is_night_shift = 'is_night_shift' in request.POST
        
        if not name:
            messages.error(request, "Shift name is required.")
            return redirect('create_shift')
        
        if not start_time or not end_time:
            messages.error(request, "Shift time is required.")
            return redirect('create_shift')

        # Check uniqueness across all records (including deleted)
        existing_shift = Shift.objects.filter(organization=request.user.organization, name__iexact=name).first()
        if existing_shift:
            if edit_shift and existing_shift.pk == edit_shift.pk:
                pass # Editing same record
            elif not existing_shift.is_deleted:
                messages.error(request, f'A shift named "{name}" already exists.')
                return redirect('create_shift')
            else:
                # Restore deleted shift
                existing_shift.is_deleted = False
                existing_shift.is_active = True
                existing_shift.grace_minutes = grace_minutes
                existing_shift.minimum_half_day_hours = minimum_half_day_hours
                existing_shift.minimum_full_day_hours = minimum_full_day_hours
                existing_shift.is_night_shift = is_night_shift
                existing_shift.updated_by = request.user
                existing_shift.save()
                messages.success(request, f'Shift "{name}" has been restored.')
                return redirect('create_shift')
        
        if edit_shift:
            edit_shift.name = name
            edit_shift.start_time = start_time
            edit_shift.end_time = end_time
            edit_shift.grace_minutes = grace_minutes
            edit_shift.minimum_half_day_hours = minimum_half_day_hours
            edit_shift.minimum_full_day_hours = minimum_full_day_hours
            edit_shift.is_night_shift = is_night_shift
            edit_shift.updated_by = request.user
            edit_shift.save()
            messages.success(request, f'Shift "{name}" updated successfully!')
        else:
            Shift.objects.create(
                organization=request.user.organization,
                name=name,
                start_time=start_time,
                end_time=end_time,
                grace_minutes=grace_minutes,
                minimum_half_day_hours=minimum_half_day_hours,
                minimum_full_day_hours=minimum_full_day_hours,
                is_night_shift=is_night_shift,
                created_by=request.user
            )
            messages.success(request, f'Shift "{name}" created successfully!')
        return redirect('create_shift')
    
    context = {
        'edit_shift': edit_shift,
        'shifts': Shift.objects.filter(organization=request.user.organization, is_deleted=False).order_by('-created_at'),
    }
    return render(request, 'createshift.html', context)

@login_required
@staff_required
def assign_shift_view(request, pk=None):
    edit_assignment = None
    
    if not pk and request.GET.get('edit'):
        pk = request.GET.get('edit')
        
    if pk:
        try:
            edit_assignment = ShiftAssignment.objects.get(pk=pk, organization=request.user.organization, is_deleted=False)
        except ShiftAssignment.DoesNotExist:
            messages.error(request, 'Shift assignment not found.')
            return redirect('assign_shift')
            
    if request.method == 'POST':
        # Handle delete
        delete_id = request.GET.get('delete')
        if delete_id:
            try:
                assignment_to_delete = ShiftAssignment.objects.get(pk=delete_id, organization=request.user.organization, is_deleted=False)
                assignment_to_delete.is_deleted = True
                assignment_to_delete.updated_by = request.user
                assignment_to_delete.save()
                messages.success(request, 'Shift assignment deleted successfully.')
            except ShiftAssignment.DoesNotExist:
                messages.error(request, 'Assignment not found.')
            return redirect('assign_shift')
            
        employee_id = request.POST.get('employee_id')
        shift_id = request.POST.get('shift_id')
        effective_from = request.POST.get('effective_from')
        effective_to = request.POST.get('effective_to') or None

        if not employee_id or not shift_id or not effective_from:
            messages.error(request, "Employee, Shift, and Effective From date are required.")
            return redirect('assign_shift')

        try:
            employee = Employee.objects.get(pk=employee_id, organization=request.user.organization, is_deleted=False)
            shift = Shift.objects.get(pk=shift_id, organization=request.user.organization, is_deleted=False)
            
            if edit_assignment:
                edit_assignment.employee = employee
                edit_assignment.shift = shift
                edit_assignment.effective_from = effective_from
                edit_assignment.effective_to = effective_to
                edit_assignment.updated_by = request.user
                
                edit_assignment.clean()
                edit_assignment.save()
                messages.success(request, f'Shift assignment for {employee.first_name} updated successfully!')
            else:
                new_assignment = ShiftAssignment(
                    organization=request.user.organization,
                    employee=employee,
                    shift=shift,
                    effective_from=effective_from,
                    effective_to=effective_to,
                    created_by=request.user
                )
                new_assignment.clean()
                new_assignment.save()
                messages.success(request, f'Shift assigned to {employee.first_name} successfully!')
                
        except (Employee.DoesNotExist, Shift.DoesNotExist):
            messages.error(request, "Selected Employee or Shift does not exist.")
        except ValidationError as e:
            msg = e.messages[0] if hasattr(e, 'messages') else str(e)
            messages.error(request, msg)
        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")
            
        return redirect('assign_shift')

    context = {
        'edit_assignment': edit_assignment,
        'employees': Employee.objects.filter(organization=request.user.organization, is_deleted=False).order_by('first_name'),
        'shifts': Shift.objects.filter(organization=request.user.organization, is_deleted=False).order_by('name'),
        'assignments': ShiftAssignment.objects.filter(organization=request.user.organization, is_deleted=False).select_related('employee', 'shift').order_by('-created_at'),
    }
    return render(request, 'assignshift.html', context)

from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from datetime import datetime
from .models import Attendance

@login_required
@require_POST
def mark_half_day_view(request, attendance_id):
    if not request.user.is_staff:
        messages.error(request, "Permission denied.")
        return redirect('attendance')
        
    try:
        att = Attendance.objects.get(id=attendance_id, organization=request.user.organization)
        
        # Override the net_work_hours to force it into the half-day bracket (e.g. 4.0 hours)
        # This will make it show as "Half Day" in the UI and payroll will give 0.5 days.
        att.net_work_hours = 4.0
        att.save()
        
        # Log the action
        from .models import AttendanceAdminAction
        AttendanceAdminAction.objects.create(
            organization=request.user.organization,
            attendance=att,
            employee=att.employee,
            performed_by=request.user,
            action_type='UPDATE',
            attendance_date=att.date,
            new_clock_in=att.clock_in,
            new_clock_out=att.clock_out,
            created_by=request.user
        )
        
        messages.success(request, f"Successfully marked half day for {att.employee.first_name}")
    except Attendance.DoesNotExist:
        messages.error(request, "Attendance record not found.")
        
    return redirect('attendance')

@login_required
def clock_in_out_view(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        # Find employee
        employee = Employee.objects.filter(email=request.user.email, organization=request.user.organization, is_active=True, is_deleted=False).first()

        if not employee:
            messages.error(request, "Your account is not linked to an active employee profile in this organization.")
            return redirect('home')

        # --- IP Restriction Check ---
        from .models import AttendanceSettings
        att_settings = AttendanceSettings.objects.filter(organization=request.user.organization).first()
        if att_settings and att_settings.network_restriction_enabled:
            allowed_ips = att_settings.allowed_ip_addresses or []
            if allowed_ips:
                x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
                if x_forwarded_for:
                    client_ip = x_forwarded_for.split(',')[0].strip()
                else:
                    client_ip = request.META.get('REMOTE_ADDR', '')
                if client_ip not in allowed_ips:
                    messages.error(
                        request,
                        "attendance check in is restricted outside the office network Please connect to your office wifi and try again."
                    )
                    return redirect('home')

        _now_local = timezone.localtime(timezone.now())
        today = _now_local.date()
        current_time = _now_local.time()

        # --- Location Restriction Check (only for clock_in) ---
        action = request.POST.get('action')
        if action == 'clock_in':
            from .models import AttendanceSettings, AttendanceLocationLog
            import math, json as _json, logging
            _log = logging.getLogger(__name__)

            att_settings_loc = AttendanceSettings.objects.filter(organization=request.user.organization).first()
            if att_settings_loc and att_settings_loc.location_restriction_enabled:
                # Require coordinates sent via POST
                raw_lat = request.POST.get('loc_lat', '').strip()
                raw_lng = request.POST.get('loc_lng', '').strip()
                raw_acc = request.POST.get('loc_accuracy', '').strip()

                if not raw_lat or not raw_lng:
                    return JsonResponse({
                        'success': False,
                        'error': 'location_required',
                        'message': 'Location permission is required to check in. Please allow location access and try again.'
                    }, status=403)

                try:
                    emp_lat = float(raw_lat)
                    emp_lng = float(raw_lng)
                    emp_acc = float(raw_acc) if raw_acc else None
                except (ValueError, TypeError):
                    return JsonResponse({
                        'success': False,
                        'error': 'invalid_coordinates',
                        'message': 'Invalid location data received. Please try again.'
                    }, status=400)

                if not (-90 <= emp_lat <= 90) or not (-180 <= emp_lng <= 180):
                    return JsonResponse({
                        'success': False,
                        'error': 'invalid_coordinates',
                        'message': 'Invalid coordinates received.'
                    }, status=400)

                # Check GPS accuracy requirement
                if emp_acc is not None and emp_acc > att_settings_loc.max_gps_accuracy_meters:
                    return JsonResponse({
                        'success': False,
                        'error': 'poor_accuracy',
                        'message': f'Your location accuracy is currently ±{int(emp_acc)} meters. '
                                   f'Attendance requires an accuracy of at least ±{att_settings_loc.max_gps_accuracy_meters} meters or better. '
                                   f'Please move to an area with better signal and try again.'
                    }, status=403)

                # Verify office coordinates are configured
                if att_settings_loc.office_latitude is None or att_settings_loc.office_longitude is None:
                    _log.error("Location restriction enabled but office coordinates not configured for org %s", request.user.organization)
                    return JsonResponse({
                        'success': False,
                        'error': 'not_configured',
                        'message': 'Office location is not configured. Please contact your administrator.'
                    }, status=503)

                # Haversine distance calculation (server-side only)
                def _haversine(lat1, lon1, lat2, lon2):
                    R = 6371000  # Earth radius in metres
                    phi1, phi2 = math.radians(lat1), math.radians(lat2)
                    dphi = math.radians(lat2 - lat1)
                    dlambda = math.radians(lon2 - lon1)
                    a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlambda/2)**2
                    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

                distance_m = _haversine(
                    emp_lat, emp_lng,
                    float(att_settings_loc.office_latitude),
                    float(att_settings_loc.office_longitude)
                )

                if distance_m > att_settings_loc.allowed_radius_meters:
                    return JsonResponse({
                        'success': False,
                        'error': 'outside_radius',
                        'message': f'You are {int(distance_m)} meters from the office. '
                                   f'Attendance check-in is only allowed within {att_settings_loc.allowed_radius_meters} meters of the office.',
                        'distance_meters': int(distance_m),
                        'allowed_meters': att_settings_loc.allowed_radius_meters,
                    }, status=403)

                # Location is valid — store audit log after attendance is saved below
                # Attach to request for use after attendance record is created
                request._loc_data = {
                    'lat': emp_lat, 'lng': emp_lng,
                    'accuracy': emp_acc or 0,
                    'distance': distance_m,
                    'verified': True,
                }


        # Get current shift assignment
        from attendance.models import ShiftAssignment
        from django.db.models import Q
        shift_assignment = ShiftAssignment.objects.filter(
            employee=employee,
            effective_from__lte=today
        ).filter(
            Q(effective_to__isnull=True) | Q(effective_to__gte=today)
        ).first()
        
        shift = shift_assignment.shift if shift_assignment else None
        
        attendance = None
        created = False
        
        if action == 'clock_out':
            attendance = Attendance.objects.filter(employee=employee, date=today, clock_out__isnull=True).first()
            if not attendance:
                yesterday = today - timedelta(days=1)
                attendance = Attendance.objects.filter(employee=employee, date=yesterday, clock_out__isnull=True, shift__is_night_shift=True).first()
                
        if not attendance:
            attendance, created = Attendance.objects.get_or_create(
                employee=employee,
                date=today,
                defaults={
                    'organization': employee.organization,
                    'shift': shift
                }
            )
        
        # If record already existed but shift was missing, update it
        if not attendance.shift and shift:
            attendance.shift = shift
            attendance.save()
        
        if action == 'clock_in':
            if attendance.clock_in:
                messages.warning(request, "You have already clocked in today.")
            else:
                attendance.clock_in = current_time
                
                # Calculate Late Minutes
                if attendance.shift:
                    shift_start = datetime.combine(attendance.date, attendance.shift.start_time)
                    actual_in = datetime.combine(today, current_time)
                    
                    if actual_in > shift_start:
                        diff = actual_in - shift_start
                        late_mins = int(diff.total_seconds() / 60)
                        if late_mins > attendance.shift.grace_minutes:
                            attendance.late_minutes = late_mins
                
                attendance.save()

                # Save location audit log if location was validated
                loc_data = getattr(request, '_loc_data', None)
                if loc_data:
                    from .models import AttendanceLocationLog
                    AttendanceLocationLog.objects.update_or_create(
                        attendance=attendance,
                        defaults={
                            'organization': employee.organization,
                            'latitude': loc_data['lat'],
                            'longitude': loc_data['lng'],
                            'location_accuracy_meters': loc_data['accuracy'],
                            'distance_from_office_meters': round(loc_data['distance'], 2),
                            'location_verified': loc_data['verified'],
                            'created_by': request.user,
                        }
                    )

                # Return JSON if the request expected it (location flow), else use messages
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.POST.get('ajax') == '1':
                    return JsonResponse({'success': True, 'message': f"Successfully clocked in at {current_time.strftime('%I:%M %p')}."})

                messages.success(request, f"Successfully clocked in at {current_time.strftime('%I:%M %p')}.")

        elif action == 'clock_out':
            if not attendance.clock_in:
                messages.warning(request, "You need to clock in first.")
            elif attendance.clock_out:
                messages.warning(request, "You have already clocked out today.")
            else:
                attendance.clock_out = current_time
                
                # Calculate Early Out Minutes
                if attendance.shift:
                    shift_end = datetime.combine(attendance.date, attendance.shift.end_time)
                    if attendance.shift.is_night_shift:
                        shift_end += timedelta(days=1)
                        
                    actual_out = datetime.combine(today, current_time)
                    
                    if actual_out < shift_end:
                        diff = shift_end - actual_out
                        attendance.early_out_minutes = int(diff.total_seconds() / 60)
                
                # Calculate Total Work Hours
                t1 = datetime.combine(attendance.date, attendance.clock_in)
                # if clocked in late night and now it's next day
                if today > attendance.date and attendance.clock_in < current_time:
                    # Actually, we know t1's true date is today - 1 (or attendance.date)
                    pass
                t2 = datetime.combine(today, current_time)
                
                if t2 < t1:
                    t2 += timedelta(days=1)
                    
                delta = t2 - t1
                attendance.total_work_hours = delta.total_seconds() / 3600.0
                
                attendance.save()
                messages.success(request, f"Successfully clocked out at {current_time.strftime('%I:%M %p')}.")
                
    return redirect('home')

@login_required
def request_correction_view(request, attendance_id):
    attendance = get_object_or_404(Attendance, id=attendance_id, employee__email=request.user.email)
    
    if request.method == 'POST':
        requested_in = request.POST.get('requested_in')
        requested_out = request.POST.get('requested_out')
        reason = request.POST.get('reason')
        
        if not reason:
            messages.error(request, "Reason is required.")
            return redirect('attendance')
            
        time_in = None
        time_out = None
        try:
            if requested_in:
                time_in = datetime.strptime(requested_in, '%H:%M').time()
            if requested_out:
                time_out = datetime.strptime(requested_out, '%H:%M').time()
        except ValueError:
            messages.error(request, "Invalid time format.")
            return redirect('attendance')
            
        AttendanceCorrection.objects.create(
            attendance=attendance,
            employee=attendance.employee,
            requested_clock_in=time_in,
            requested_clock_out=time_out,
            reason=reason
        )
        messages.success(request, "Correction request submitted successfully.")
    
    return redirect('attendance')

@login_required
def manage_corrections_view(request):
    if not request.user.is_staff:
        messages.error(request, "Access denied.")
        return redirect('home')
        
    corrections = AttendanceCorrection.objects.filter(employee__organization=request.user.organization, status='PENDING').order_by('-created_at')
    
    context = {
        'corrections': corrections
    }
    return render(request, 'manage_corrections.html', context)

@login_required
def resolve_correction_view(request, correction_id):
    if not request.user.is_staff:
        messages.error(request, "Access denied.")
        return redirect('home')
        
    correction = get_object_or_404(AttendanceCorrection, id=correction_id, employee__organization=request.user.organization)
    
    if request.method == 'POST':
        from django.utils import timezone
        action = request.POST.get('action')
        now = timezone.now()
        
        if action == 'APPROVE':
            correction.status = 'APPROVED'
            correction.resolved_by = request.user
            correction.resolved_at = now
            correction.save()
            
            att = correction.attendance
            if correction.requested_clock_in:
                att.clock_in = correction.requested_clock_in
            if correction.requested_clock_out:
                att.clock_out = correction.requested_clock_out
                
            if att.shift and att.clock_in:
                shift_start = datetime.combine(att.date, att.shift.start_time)
                actual_in = datetime.combine(att.date, att.clock_in)
                # If night shift and clock in is early morning
                if att.shift.is_night_shift and att.clock_in < att.shift.start_time and att.clock_in < time(12,0):
                    actual_in += timedelta(days=1)
                    
                if actual_in > shift_start:
                    diff = actual_in - shift_start
                    late_mins = int(diff.total_seconds() / 60)
                    if late_mins > att.shift.grace_minutes:
                        att.late_minutes = late_mins
                    else:
                        att.late_minutes = 0
                else:
                    att.late_minutes = 0
                    
            if att.shift and att.clock_out:
                shift_end = datetime.combine(att.date, att.shift.end_time)
                if att.shift.is_night_shift:
                    shift_end += timedelta(days=1)
                    
                actual_out = datetime.combine(att.date, att.clock_out)
                if att.shift.is_night_shift and att.clock_out < att.shift.start_time:
                    actual_out += timedelta(days=1)
                    
                if actual_out < shift_end:
                    diff = shift_end - actual_out
                    att.early_out_minutes = int(diff.total_seconds() / 60)
                else:
                    att.early_out_minutes = 0
                    
            if att.clock_in and att.clock_out:
                t1 = datetime.combine(att.date, att.clock_in)
                if att.shift and att.shift.is_night_shift and att.clock_in < att.shift.start_time and att.clock_in < time(12,0):
                    t1 += timedelta(days=1)
                    
                t2 = datetime.combine(att.date, att.clock_out)
                if t2 < t1:
                    t2 += timedelta(days=1)
                    
                delta = t2 - t1
                att.total_work_hours = delta.total_seconds() / 3600.0
                
            att.save()
            messages.success(request, "Correction approved and attendance updated.")
            
        elif action == 'REJECT':
            correction.status = 'REJECTED'
            correction.resolved_by = request.user
            correction.resolved_at = now
            correction.save()
            messages.success(request, "Correction request rejected.")
            
    return redirect('manage_corrections')

@login_required
def break_toggle_view(request):
    from django.utils import timezone
    today = timezone.now().date()
    employee = Employee.objects.filter(email=request.user.email, is_active=True, is_deleted=False).first()
    if not employee:
        messages.error(request, "Employee record not found.")
        return redirect('home')
        
    attendance = Attendance.objects.filter(employee=employee, date=today).first()
    if not attendance or not attendance.clock_in:
        messages.error(request, "You must be clocked in to take a break.")
        return redirect('home')
        
    if attendance.clock_out:
        messages.error(request, "You have already clocked out for today.")
        return redirect('home')
        
    from .models import BreakLog
    active_break = BreakLog.objects.filter(attendance=attendance, end_time__isnull=True).first()
    
    if active_break:
        now = timezone.now()
        active_break.end_time = now
        delta = now - active_break.start_time
        active_break.duration_minutes = int(delta.total_seconds() / 60)
        active_break.save()
        
        all_breaks = BreakLog.objects.filter(attendance=attendance, end_time__isnull=False)
        total_mins = sum(b.duration_minutes for b in all_breaks)
        attendance.total_break_minutes = total_mins
        
        if attendance.clock_in:
            t1 = timezone.make_aware(datetime.combine(attendance.date, attendance.clock_in))
            t2 = now
            gross_seconds = (t2 - t1).total_seconds()
            net_seconds = gross_seconds - (total_mins * 60)
            attendance.net_work_hours = max(0, net_seconds / 3600.0)
            
        attendance.save()
        messages.success(request, f"Break ended. Duration: {active_break.duration_minutes} mins.")
    else:
        BreakLog.objects.create(attendance=attendance)
        messages.success(request, "Break started. Stay refreshed!")
        
    return redirect('home')

@login_required
def attendance_calendar_view(request):
    organization = request.user.organization
    
    # Staff can view any employee's calendar, others only their own
    employee_id = request.GET.get('employee_id')
    if request.user.is_staff and employee_id:
        employee = get_object_or_404(Employee, id=employee_id, organization=organization)
    else:
        employee = Employee.objects.filter(email=request.user.email, organization=organization, is_active=True, is_deleted=False).first()
    
    if not employee:
        messages.error(request, "Employee profile not found.")
        return redirect('home')

    # Get month and year from GET parameters or use current
    today = timezone.now().date()
    year = int(request.GET.get('year', today.year))
    month = int(request.GET.get('month', today.month))
    
    # Calculate prev/next month
    first_day = date(year, month, 1)
    prev_month_date = first_day - timedelta(days=1)
    next_month_date = (first_day + timedelta(days=32)).replace(day=1)
    
    # Get all attendance for this employee in this month
    attendances = Attendance.objects.filter(
        employee=employee,
        date__year=year,
        date__month=month
    )
    attendance_map = {att.date: att for att in attendances}
    
    # Get holidays
    from leaves.models import Holiday
    holidays = Holiday.objects.filter(
        organization=organization,
        date__year=year,
        date__month=month
    )
    holiday_map = {h.date: h for h in holidays}
    
    # Build calendar
    cal = calendar.Calendar(firstweekday=6) # Sunday start
    month_days = cal.monthdays2calendar(year, month)
    
    # Process days for template
    processed_calendar = []
    for week in month_days:
        week_days = []
        for day_num, weekday in week:
            if day_num == 0:
                week_days.append({'day': 0, 'status': 'empty'})
            else:
                curr_date = date(year, month, day_num)
                att = attendance_map.get(curr_date)
                hol = holiday_map.get(curr_date)
                
                status = 'none'
                label = ''
                
                if hol:
                    status = 'holiday'
                    label = hol.name
                elif att:
                    if att.late_minutes > 0:
                        status = 'late'
                    elif att.clock_in:
                        status = 'present'
                    else:
                        status = 'absent'
                elif curr_date < today:
                    # If past date and no attendance/holiday, mark as absent if it's a weekday
                    if curr_date.weekday() != 6:  # Mon–Sat (Sunday only is weekly off)
                         status = 'absent'
                    else:
                         status = 'weekend'
                
                week_days.append({
                    'day': day_num,
                    'date': curr_date,
                    'status': status,
                    'label': label,
                    'attendance': att
                })
        processed_calendar.append(week_days)
    
    # For staff, get list of employees for selector
    employees = None
    if request.user.is_staff:
        employees = Employee.objects.filter(organization=organization, is_deleted=False).order_by('first_name')
        
    context = {
        'calendar': processed_calendar,
        'year': year,
        'month': month,
        'month_name': calendar.month_name[month],
        'prev_year': prev_month_date.year,
        'prev_month': prev_month_date.month,
        'next_year': next_month_date.year,
        'next_month': next_month_date.month,
        'today': today,
        'employee': employee,
        'employees': employees,
    }
    
    return render(request, 'attendance_calendar.html', context)

@login_required
def export_attendance_view(request):
    if not request.user.is_staff:
        messages.error(request, "Access denied.")
        return redirect('home')
        
    organization = request.user.organization
    export_range = request.GET.get('range', 'today')
    export_format = request.GET.get('format', 'excel')
    
    today = timezone.now().date()
    start_date = today
    end_date = today
    
    if export_range == 'today':
        start_date = today
        end_date = today
    elif export_range == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = today
    elif export_range == 'month':
        start_date = today.replace(day=1)
        end_date = today
    elif export_range == 'custom':
        s_str = request.GET.get('start_date')
        e_str = request.GET.get('end_date')
        try:
            start_date = datetime.strptime(s_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(e_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            messages.error(request, "Invalid date range.")
            return redirect('attendance')
            
    attendances = Attendance.objects.filter(
        employee__organization=organization, 
        date__range=[start_date, end_date]
    ).select_related('employee', 'employee__department').order_by('-date', 'employee__first_name')
    
    import io
    from django.http import HttpResponse

    if export_format == 'excel':
        import xlsxwriter
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet("Attendance Report")
        
        header_fmt = workbook.add_format({'bold': True, 'bg_color': '#4F46E5', 'font_color': 'white', 'border': 1})
        border_fmt = workbook.add_format({'border': 1})
        
        headers = ['Date', 'Employee ID', 'Name', 'Department', 'Clock In', 'Clock Out', 'Gross Hours', 'Break (M)', 'Net Hours', 'Status']
        for col, text in enumerate(headers):
            worksheet.write(0, col, text, header_fmt)
            
        for row, log in enumerate(attendances, start=1):
            worksheet.write(row, 0, log.date.strftime('%Y-%m-%d'), border_fmt)
            worksheet.write(row, 1, log.employee.employee_id, border_fmt)
            worksheet.write(row, 2, f"{log.employee.first_name} {log.employee.last_name}", border_fmt)
            worksheet.write(row, 3, log.employee.department.name if log.employee.department else "--", border_fmt)
            worksheet.write(row, 4, log.clock_in.strftime('%I:%M %p') if log.clock_in else "--", border_fmt)
            worksheet.write(row, 5, log.clock_out.strftime('%I:%M %p') if log.clock_out else "--", border_fmt)
            worksheet.write(row, 6, log.current_work_time, border_fmt)
            worksheet.write(row, 7, log.total_break_minutes or 0, border_fmt)
            worksheet.write(row, 8, f"{log.net_work_hours:.2f}h" if log.net_work_hours else log.current_work_time, border_fmt)
            
            status = "Absent"
            if log.clock_in and log.clock_out: status = "Completed"
            elif log.clock_in: status = "Working"
            worksheet.write(row, 9, status, border_fmt)
            
        worksheet.set_column(0, 9, 15)
        workbook.close()
        output.seek(0)
        
        response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = f'attachment; filename="Attendance_{start_date}_to_{end_date}.xlsx"'
        return response

    elif export_format == 'pdf':
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4))
        elements = []
        styles = getSampleStyleSheet()
        
        elements.append(Paragraph(f"Attendance Report: {start_date} to {end_date}", styles['Title']))
        elements.append(Spacer(1, 12))
        
        data = [['Date', 'Emp ID', 'Name', 'Dept', 'In', 'Out', 'Gross', 'Break', 'Net', 'Status']]
        for log in attendances:
            status = "Absent"
            if log.clock_in and log.clock_out: status = "Completed"
            elif log.clock_in: status = "Working"
            
            data.append([
                log.date.strftime('%d/%m'),
                log.employee.employee_id,
                f"{log.employee.first_name} {log.employee.last_name}"[:20],
                log.employee.department.name[:10] if log.employee.department else "--",
                log.clock_in.strftime('%I:%M%p') if log.clock_in else "--",
                log.clock_out.strftime('%I:%M%p') if log.clock_out else "--",
                log.current_work_time,
                str(log.total_break_minutes or 0),
                f"{log.net_work_hours:.2f}h" if log.net_work_hours else log.current_work_time,
                status
            ])
            
        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(t)
        doc.build(elements)
        
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Attendance_{start_date}_to_{end_date}.pdf"'
        return response

    return redirect('attendance')

from django.db.models import Avg, Sum, Count, Q
from django.utils import timezone
from datetime import timedelta

@login_required
def attendance_analytics_view(request):
    if not request.user.is_staff:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("You are not authorized to view this page.")
        
    today = timezone.localdate()
    start_of_month = today.replace(day=1)
    
    month_attendances = Attendance.objects.filter(date__gte=start_of_month, date__lte=today)
    
    from employees.models import Employee, Department
    total_employees = Employee.objects.count()
    
    today_attendances = Attendance.objects.filter(date=today)
    present_today = today_attendances.filter(clock_in__isnull=False).count()
    absent_today = total_employees - present_today
    
    total_punches = month_attendances.filter(clock_in__isnull=False).count()
    late_punches = month_attendances.filter(late_minutes__gt=0).count()
    early_departures = month_attendances.filter(early_out_minutes__gt=0).count()
    on_time_punches = total_punches - late_punches
    
    punctuality_percent = 0
    if total_punches > 0:
        punctuality_percent = round((on_time_punches / total_punches) * 100, 1)
        
    avg_hours_decimal = month_attendances.aggregate(avg=Avg('net_work_hours'))['avg']
    avg_hours = round(avg_hours_decimal, 1) if avg_hours_decimal else 0
    
    # 7-Day Trend
    last_7_days = [today - timedelta(days=i) for i in range(6, -1, -1)]
    trend_labels = [d.strftime('%b %d') for d in last_7_days]
    
    trend_present = []
    trend_late = []
    trend_overtime = []
    
    for d in last_7_days:
        day_att = Attendance.objects.filter(date=d)
        p_count = day_att.filter(clock_in__isnull=False).count()
        l_count = day_att.filter(late_minutes__gt=0).count()
        o_hours = day_att.aggregate(total=Sum('overtime_hours'))['total'] or 0
        trend_present.append(p_count)
        trend_late.append(l_count)
        trend_overtime.append(float(o_hours))
        
    # Department Wise Attendance
    departments = Department.objects.filter(organization=request.user.organization, is_deleted=False)
    dept_labels = []
    dept_present_rates = []
    
    for dept in departments:
        dept_emps = Employee.objects.filter(department=dept, is_active=True).count()
        if dept_emps > 0:
            dept_att = today_attendances.filter(employee__department=dept, clock_in__isnull=False).count()
            rate = round((dept_att / dept_emps) * 100, 1)
            dept_labels.append(dept.name)
            dept_present_rates.append(rate)

    # Top Latecomers
    late_list = month_attendances.filter(late_minutes__gt=0).values('employee__first_name', 'employee__last_name').annotate(total_late=Sum('late_minutes'), times_late=Count('id')).order_by('-total_late')[:5]

    context = {
        'total_employees': total_employees,
        'present_today': present_today,
        'absent_today': absent_today,
        'punctuality_percent': punctuality_percent,
        'avg_hours': avg_hours,
        'early_departures': early_departures,
        'trend_labels': trend_labels,
        'trend_present': trend_present,
        'trend_late': trend_late,
        'trend_overtime': trend_overtime,
        'dept_labels': dept_labels,
        'dept_present_rates': dept_present_rates,
        'late_list': late_list,
    }
    
    return render(request, 'attendance_analytics.html', context)

@login_required
def overtime_dashboard_view(request):
    from employees.models import Employee
    from .models import OvertimeRequest
    
    employee = Employee.objects.filter(email=request.user.email, is_active=True, is_deleted=False).first()
    
    if request.method == 'POST' and employee:
        date = request.POST.get('date')
        hours = request.POST.get('hours')
        reason = request.POST.get('reason')
        
        if date and hours and reason:
            OvertimeRequest.objects.create(
                employee=employee,
                date=date,
                hours_requested=hours,
                reason=reason,
                created_by=request.user
            )
            messages.success(request, "Overtime request submitted successfully.")
            return redirect('overtime_dashboard')
            
    if employee and not request.user.is_staff:
        requests_list = OvertimeRequest.objects.filter(employee=employee).order_by('-created_at')
        pending_requests = None
        
    elif request.user.is_staff:
        requests_list = OvertimeRequest.objects.filter(employee__organization=request.user.organization).order_by('-created_at')
        pending_requests = OvertimeRequest.objects.filter(employee__organization=request.user.organization, status='PENDING').order_by('-created_at')
        
    else:
        requests_list = []
        pending_requests = None
        
    context = {
        'requests': requests_list,
        'pending_requests': pending_requests,
        'employee': employee,
    }
    
    return render(request, 'overtime_dashboard.html', context)

@login_required
def overtime_action_view(request, request_id):
    if not request.user.is_staff:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("Access Denied")
        
    from .models import OvertimeRequest
    from django.shortcuts import get_object_or_404
    
    ot_request = get_object_or_404(OvertimeRequest, id=request_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'APPROVE':
            ot_request.status = 'APPROVED'
            ot_request.approved_by = request.user
            ot_request.save()
            messages.success(request, "Overtime approved.")
            
            att = Attendance.objects.filter(employee=ot_request.employee, date=ot_request.date).first()
            if att:
                att.overtime_hours = float(att.overtime_hours) + float(ot_request.hours_requested)
                att.save()
                
        elif action == 'REJECT':
            ot_request.status = 'REJECTED'
            ot_request.approved_by = request.user
            ot_request.save()
            messages.success(request, "Overtime rejected.")
            
    return redirect('overtime_dashboard')


# ─────────────────────────────────────────────────────────────────────────────
# ADMIN ATTENDANCE OVERRIDE (Create / Edit)
# ─────────────────────────────────────────────────────────────────────────────

from .models import AttendanceAdminAction
from django.views.decorators.http import require_POST
from django.shortcuts import get_object_or_404 as _get_object_or_404


def _parse_time(value):
    """Parse 'HH:MM' or 'HH:MM:SS' string → datetime.time or None."""
    if not value:
        return None
    try:
        parts = value.split(':')
        return time(int(parts[0]), int(parts[1]))
    except (ValueError, IndexError, AttributeError):
        return None


@login_required
@require_POST
def admin_attendance_create_view(request, employee_id):
    """Admin creates a fresh attendance record for an employee on a given date."""
    if not request.user.is_staff:
        return JsonResponse({'ok': False, 'error': 'Permission denied.'}, status=403)

    org = request.user.organization
    employee = _get_object_or_404(Employee, pk=employee_id, organization=org, is_active=True, is_deleted=False)

    date_str    = request.POST.get('attendance_date', '')
    clock_in_s  = request.POST.get('clock_in', '').strip()
    clock_out_s = request.POST.get('clock_out', '').strip()
    late_min_s  = request.POST.get('late_minutes', '0').strip()

    # --- Validate date ---
    try:
        att_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'ok': False, 'error': 'Invalid date.'}, status=400)

    # --- Parse times ---
    clock_in  = _parse_time(clock_in_s)
    clock_out = _parse_time(clock_out_s) if clock_out_s else None

    if not clock_in:
        return JsonResponse({'ok': False, 'error': 'Check-in time is required.'}, status=400)

    if clock_out and clock_out <= clock_in:
        return JsonResponse({'ok': False, 'error': 'Check-out cannot be earlier than check-in.'}, status=400)

    try:
        late_minutes = int(late_min_s)
        if late_minutes < 0:
            raise ValueError
    except ValueError:
        return JsonResponse({'ok': False, 'error': 'Late minutes must be a non-negative integer.'}, status=400)

    # --- Guard: must not already have a record ---
    if Attendance.objects.filter(employee=employee, date=att_date).exists():
        return JsonResponse({'ok': False, 'error': 'Attendance already exists for this date. Use Edit instead.'}, status=400)

    with transaction.atomic():
        att = Attendance.objects.create(
            organization=org,
            employee=employee,
            date=att_date,
            clock_in=clock_in,
            clock_out=clock_out,
            late_minutes=late_minutes,
            created_by=request.user,
        )
        AttendanceAdminAction.objects.create(
            organization=org,
            attendance=att,
            employee=employee,
            performed_by=request.user,
            action_type='CREATE',
            attendance_date=att_date,
            old_clock_in=None,
            old_clock_out=None,
            old_late_minutes=None,
            new_clock_in=clock_in,
            new_clock_out=clock_out,
            new_late_minutes=late_minutes,
            created_by=request.user,
        )

    return JsonResponse({
        'ok': True,
        'message': 'Attendance created successfully.',
        'clock_in': clock_in.strftime('%I:%M %p') if clock_in else '—',
        'clock_out': clock_out.strftime('%I:%M %p') if clock_out else '—',
        'late_minutes': late_minutes,
    })


@login_required
@require_POST
def admin_attendance_edit_view(request, attendance_id):
    """Admin edits an existing attendance record."""
    if not request.user.is_staff:
        return JsonResponse({'ok': False, 'error': 'Permission denied.'}, status=403)

    org = request.user.organization
    att = _get_object_or_404(Attendance, pk=attendance_id, employee__organization=org)

    clock_in_s  = request.POST.get('clock_in', '').strip()
    clock_out_s = request.POST.get('clock_out', '').strip()
    late_min_s  = request.POST.get('late_minutes', '0').strip()

    clock_in  = _parse_time(clock_in_s)
    clock_out = _parse_time(clock_out_s) if clock_out_s else None

    if not clock_in:
        return JsonResponse({'ok': False, 'error': 'Check-in time is required.'}, status=400)

    if clock_out and clock_out <= clock_in:
        return JsonResponse({'ok': False, 'error': 'Check-out cannot be earlier than check-in.'}, status=400)

    try:
        late_minutes = int(late_min_s)
        if late_minutes < 0:
            raise ValueError
    except ValueError:
        return JsonResponse({'ok': False, 'error': 'Late minutes must be a non-negative integer.'}, status=400)

    with transaction.atomic():
        # Snapshot old values
        old_ci   = att.clock_in
        old_co   = att.clock_out
        old_late = att.late_minutes

        att.clock_in      = clock_in
        att.clock_out     = clock_out
        att.late_minutes  = late_minutes
        att.updated_by    = request.user
        att.save()

        AttendanceAdminAction.objects.create(
            organization=org,
            attendance=att,
            employee=att.employee,
            performed_by=request.user,
            action_type='UPDATE',
            attendance_date=att.date,
            old_clock_in=old_ci,
            old_clock_out=old_co,
            old_late_minutes=old_late,
            new_clock_in=clock_in,
            new_clock_out=clock_out,
            new_late_minutes=late_minutes,
            created_by=request.user,
        )

    return JsonResponse({
        'ok': True,
        'message': 'Attendance updated successfully.',
        'clock_in': clock_in.strftime('%I:%M %p') if clock_in else '—',
        'clock_out': clock_out.strftime('%I:%M %p') if clock_out else '—',
        'late_minutes': late_minutes,
    })


@login_required
def admin_attendance_history_view(request, attendance_id):
    """Return audit history for a given attendance record (staff only)."""
    if not request.user.is_staff:
        return JsonResponse({'ok': False, 'error': 'Permission denied.'}, status=403)

    org = request.user.organization
    att = _get_object_or_404(Attendance, pk=attendance_id, employee__organization=org)

    logs = AttendanceAdminAction.objects.filter(attendance=att).select_related('performed_by').order_by('-created_at')

    def fmt_t(t):
        return t.strftime('%I:%M %p') if t else '—'

    history = []
    for log in logs:
        history.append({
            'action_type': log.action_type,
            'performed_by': f"{log.performed_by.get_full_name() or log.performed_by.username}" if log.performed_by else '—',
            'is_staff': log.performed_by.is_staff if log.performed_by else False,
            'created_at': log.created_at.strftime('%d %b %Y · %I:%M %p'),
            'old_clock_in':  fmt_t(log.old_clock_in),
            'old_clock_out': fmt_t(log.old_clock_out),
            'old_late_minutes': log.old_late_minutes,
            'new_clock_in':  fmt_t(log.new_clock_in),
            'new_clock_out': fmt_t(log.new_clock_out),
            'new_late_minutes': log.new_late_minutes,
        })

    return JsonResponse({'ok': True, 'history': history, 'employee': str(att.employee), 'date': att.date.strftime('%d %b %Y')})


# ─────────────────────────────────────────────────────────────────────────────
# BULK EXCEL EXPORT / IMPORT
# ─────────────────────────────────────────────────────────────────────────────

import calendar as _cal_module
import io


@login_required
def export_monthly_template_view(request):
    """Generate and download a pre-filled monthly attendance Excel template."""
    if not request.user.is_staff:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("Permission denied.")

    try:
        month = int(request.GET.get('month', timezone.localdate().month))
        year  = int(request.GET.get('year',  timezone.localdate().year))
    except ValueError:
        messages.error(request, "Invalid month/year.")
        return redirect('attendance')

    if not (1 <= month <= 12):
        messages.error(request, "Month must be between 1 and 12.")
        return redirect('attendance')

    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, Protection
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from leaves.models import Holiday

    org          = request.user.organization
    month_start  = date(year, month, 1)
    month_end    = date(year, month, _cal_module.monthrange(year, month)[1])
    month_days   = [month_start + timedelta(days=i) for i in range((month_end - month_start).days + 1)]

    # Fetch all active employees
    employees = list(
        Employee.objects.filter(organization=org, is_active=True, is_deleted=False)
        .select_related('department')
        .order_by('first_name', 'last_name', 'employee_id')
    )

    # Fetch holidays
    holidays = Holiday.objects.filter(
        organization=org,
        date__range=(month_start, month_end),
        is_deleted=False,
    )
    holiday_dates = {h.date: h.name for h in holidays}

    # ── Styles ──────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # ── Attendance Sheet ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Attendance"

    BLUE_FILL   = PatternFill("solid", fgColor="2563EB")
    GRAY_FILL   = PatternFill("solid", fgColor="F1F5F9")
    WKND_FILL   = PatternFill("solid", fgColor="CBD5E1")
    HOLI_FILL   = PatternFill("solid", fgColor="DBEAFE")
    ALT_FILL    = PatternFill("solid", fgColor="F8FAFC")
    WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")

    HDR_FONT    = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    SUBHDR_FONT = Font(name="Calibri", bold=True, color="1E3A5F", size=9)
    EMP_FONT    = Font(name="Calibri", bold=True, color="0F172A", size=9)
    CELL_FONT   = Font(name="Calibri", size=9)
    WKND_FONT   = Font(name="Calibri", color="64748B", size=9, italic=True)
    HOLI_FONT   = Font(name="Calibri", color="1D4ED8", size=9, italic=True)

    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left",   vertical="center")

    thin = Side(style="thin", color="E2E8F0")
    med  = Side(style="medium", color="CBD5E1")
    THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    MED_BORDER  = Border(left=med,  right=med,  top=med,  bottom=med)

    # Row 1 – title
    ws.row_dimensions[1].height = 22
    ws.merge_cells(f"A1:{get_column_letter(3 + len(month_days) * 3)}1")
    title_cell = ws["A1"]
    title_cell.value       = f"Monthly Attendance Template — {month_start.strftime('%B %Y')}  |  {org.name}"
    title_cell.font        = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    title_cell.fill        = BLUE_FILL
    title_cell.alignment   = CENTER

    # Row 2 – date group headers (merged per date triplet)
    ws.row_dimensions[2].height = 28
    ws["A2"].value = "Employee ID";  ws["A2"].font = SUBHDR_FONT; ws["A2"].fill = GRAY_FILL; ws["A2"].alignment = CENTER; ws["A2"].border = THIN_BORDER
    ws["B2"].value = "Employee Name"; ws["B2"].font = SUBHDR_FONT; ws["B2"].fill = GRAY_FILL; ws["B2"].alignment = LEFT;   ws["B2"].border = THIN_BORDER
    ws["C2"].value = "Department";    ws["C2"].font = SUBHDR_FONT; ws["C2"].fill = GRAY_FILL; ws["C2"].alignment = LEFT;   ws["C2"].border = THIN_BORDER

    for idx, d in enumerate(month_days):
        base_col = 4 + idx * 3          # 1-indexed
        is_weekend = (d.weekday() == 6)  # Sunday
        is_holiday = d in holiday_dates
        hdr_fill = HOLI_FILL if is_holiday else (WKND_FILL if is_weekend else GRAY_FILL)
        hdr_font = HOLI_FONT if is_holiday else (WKND_FONT if is_weekend else SUBHDR_FONT)
        lbl = d.strftime("%d %b")
        if is_holiday:
            lbl += f"\n({holiday_dates[d][:12]})"
        elif is_weekend:
            lbl += "\n(Sun)"

        merge_start = get_column_letter(base_col)
        merge_end   = get_column_letter(base_col + 2)
        ws.merge_cells(f"{merge_start}2:{merge_end}2")
        cell = ws[f"{merge_start}2"]
        cell.value     = lbl
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = CENTER
        cell.border    = MED_BORDER

    # Row 3 – sub-headers: Status | Clock In | Clock Out per day
    ws.row_dimensions[3].height = 18
    ws["A3"].value = ""; ws["A3"].fill = GRAY_FILL; ws["A3"].border = THIN_BORDER
    ws["B3"].value = ""; ws["B3"].fill = GRAY_FILL; ws["B3"].border = THIN_BORDER
    ws["C3"].value = ""; ws["C3"].fill = GRAY_FILL; ws["C3"].border = THIN_BORDER
    for idx, d in enumerate(month_days):
        base_col = 4 + idx * 3
        is_weekend = (d.weekday() == 6)
        is_holiday = d in holiday_dates
        hdr_fill = HOLI_FILL if is_holiday else (WKND_FILL if is_weekend else GRAY_FILL)
        for offset, label in enumerate(["Status", "Clock In", "Clock Out"]):
            c = ws.cell(row=3, column=base_col + offset)
            c.value     = label
            c.font      = SUBHDR_FONT
            c.fill      = hdr_fill
            c.alignment = CENTER
            c.border    = THIN_BORDER

    # ── Dropdown validation for Status (working days only) ─────────────────
    valid_statuses = '"Present,Absent,Half Day,Leave"'
    dv = DataValidation(type="list", formula1=valid_statuses, allow_blank=True, showDropDown=False)
    dv.error       = "Invalid status. Use: Present, Absent, Half Day, Leave"
    dv.errorTitle  = "Invalid Value"
    dv.prompt      = "Select attendance status"
    dv.promptTitle = "Status"
    ws.add_data_validation(dv)

    # ── Employee rows ────────────────────────────────────────────────────────
    for emp_idx, emp in enumerate(employees):
        row = 4 + emp_idx
        ws.row_dimensions[row].height = 16
        row_fill = ALT_FILL if emp_idx % 2 == 0 else WHITE_FILL

        # Employee ID
        c = ws.cell(row=row, column=1, value=emp.employee_id)
        c.font = EMP_FONT; c.alignment = CENTER; c.fill = row_fill; c.border = THIN_BORDER

        # Name
        c = ws.cell(row=row, column=2, value=f"{emp.first_name} {emp.last_name}")
        c.font = EMP_FONT; c.alignment = LEFT; c.fill = row_fill; c.border = THIN_BORDER

        # Department
        c = ws.cell(row=row, column=3, value=emp.department.name if emp.department else "")
        c.font = EMP_FONT; c.alignment = LEFT; c.fill = row_fill; c.border = THIN_BORDER

        for idx, d in enumerate(month_days):
            base_col   = 4 + idx * 3
            is_weekend = (d.weekday() == 6)
            is_holiday = d in holiday_dates

            # Check joining date
            if emp.joining_date and d < emp.joining_date:
                status_val = "N/A"
                protect    = True
                s_fill     = GRAY_FILL
                s_font     = WKND_FONT
            elif is_holiday:
                status_val = "HOLIDAY"
                protect    = True
                s_fill     = HOLI_FILL
                s_font     = HOLI_FONT
            elif is_weekend:
                status_val = "WEEKOFF"
                protect    = True
                s_fill     = WKND_FILL
                s_font     = WKND_FONT
            else:
                status_val = ""
                protect    = False
                s_fill     = row_fill
                s_font     = CELL_FONT

            # Status cell
            sc = ws.cell(row=row, column=base_col, value=status_val)
            sc.font = s_font; sc.fill = s_fill; sc.alignment = CENTER; sc.border = THIN_BORDER
            if not protect:
                dv.add(sc)

            # Clock In
            ci = ws.cell(row=row, column=base_col + 1, value="" if protect else "")
            ci.font = CELL_FONT; ci.fill = s_fill if protect else row_fill
            ci.alignment = CENTER; ci.border = THIN_BORDER
            ci.number_format = "HH:MM"

            # Clock Out
            co = ws.cell(row=row, column=base_col + 2, value="" if protect else "")
            co.font = CELL_FONT; co.fill = s_fill if protect else row_fill
            co.alignment = CENTER; co.border = THIN_BORDER
            co.number_format = "HH:MM"

    # ── Column widths ────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    for idx in range(len(month_days)):
        base_col = 4 + idx * 3
        ws.column_dimensions[get_column_letter(base_col)].width     = 10  # Status
        ws.column_dimensions[get_column_letter(base_col + 1)].width = 9   # Clock In
        ws.column_dimensions[get_column_letter(base_col + 2)].width = 9   # Clock Out

    # Freeze panes: freeze first 3 rows and first 3 columns
    ws.freeze_panes = "D4"

    # ── Instructions Sheet ───────────────────────────────────────────────────
    ws_instr = wb.create_sheet("Instructions")
    ws_instr.sheet_view.showGridLines = False
    instructions = [
        ("MULZON HRMS — Attendance Import Instructions", True, 14, "2563EB"),
        ("", False, 10, None),
        ("HOW TO USE THIS TEMPLATE", True, 11, "1E3A5F"),
        ("1. Do NOT rename or rearrange columns.", False, 10, None),
        ("2. Do NOT change Employee IDs — they are used to identify employees during import.", False, 10, None),
        ("3. Do NOT modify HOLIDAY or WEEKOFF cells — they are auto-filled.", False, 10, None),
        ("4. Fill only Status, Clock In, and Clock Out for working days.", False, 10, None),
        ("5. Upload the completed file via: Attendance → Import Attendance", False, 10, None),
        ("", False, 10, None),
        ("ACCEPTED STATUS VALUES", True, 11, "1E3A5F"),
        ("Present   — Employee was present.", False, 10, None),
        ("Absent    — Employee was absent (leave Clock In & Clock Out empty).", False, 10, None),
        ("Half Day  — Employee worked a half day.", False, 10, None),
        ("Leave     — Employee was on approved leave.", False, 10, None),
        ("HOLIDAY   — Auto-filled. Do not modify.", False, 10, None),
        ("WEEKOFF   — Auto-filled (Sunday). Do not modify.", False, 10, None),
        ("", False, 10, None),
        ("TIME FORMAT", True, 11, "1E3A5F"),
        ("Use 24-hour format:  HH:MM   e.g.  09:30  or  18:15", False, 10, None),
        ("", False, 10, None),
        ("RULES", True, 11, "1E3A5F"),
        ("• Clock Out cannot be earlier than Clock In.", False, 10, None),
        ("• For Absent status, leave Clock In and Clock Out empty.", False, 10, None),
        ("• Duplicate records will be flagged during import — you can choose to Skip or Update.", False, 10, None),
        ("• All changes are recorded in the audit log.", False, 10, None),
    ]
    for r, (text, bold, size, color) in enumerate(instructions, start=1):
        cell = ws_instr.cell(row=r, column=1, value=text)
        cell.font = Font(
            name="Calibri", bold=bold, size=size,
            color=color if color else "0F172A"
        )
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws_instr.column_dimensions["A"].width = 80

    # ── Build response ───────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"Attendance_Template_{month_start.strftime('%B_%Y')}_{org.name.replace(' ','_')}.xlsx"
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{fname}"'
    return response


# ─── IMPORT: Parse & Preview ─────────────────────────────────────────────────

def _parse_excel_time(raw):
    """Convert an openpyxl cell value to datetime.time or None."""
    if raw is None or raw == "":
        return None
    import datetime as _dt
    if isinstance(raw, _dt.time):
        return raw
    if isinstance(raw, _dt.datetime):
        return raw.time()
    if isinstance(raw, float):
        # Excel stores time as fraction of a day
        total_sec = round(raw * 86400)
        h, rem    = divmod(total_sec, 3600)
        m, _      = divmod(rem, 60)
        return _dt.time(h % 24, m)
    s = str(raw).strip()
    for fmt in ("%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M%p"):
        try:
            return _dt.datetime.strptime(s, fmt).time()
        except ValueError:
            pass
    return None


@login_required
def import_attendance_view(request):
    """Show upload form; on POST validate file and store parsed data in session for preview."""
    if not request.user.is_staff:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("Permission denied.")

    if request.method == "GET":
        today = timezone.localdate()
        return render(request, "attendance_import.html", {
            "current_month": today.month,
            "current_year":  today.year,
            "years": range(today.year - 2, today.year + 1),
        })

    # POST — parse & validate
    uploaded = request.FILES.get("excel_file")
    if not uploaded:
        messages.error(request, "Please upload an Excel file.")
        return redirect("import_attendance")

    # File type check
    if not uploaded.name.lower().endswith((".xlsx", ".xls")):
        messages.error(request, "Only .xlsx files are accepted.")
        return redirect("import_attendance")

    if uploaded.size > 20 * 1024 * 1024:  # 20 MB
        messages.error(request, "File too large. Maximum allowed size is 20 MB.")
        return redirect("import_attendance")

    try:
        month = int(request.POST.get("month", 0))
        year  = int(request.POST.get("year",  0))
    except ValueError:
        messages.error(request, "Invalid month/year.")
        return redirect("import_attendance")

    if not (1 <= month <= 12):
        messages.error(request, "Invalid month.")
        return redirect("import_attendance")

    import openpyxl
    from leaves.models import Holiday

    try:
        wb = openpyxl.load_workbook(uploaded, data_only=True)
    except Exception:
        messages.error(request, "Cannot read the Excel file. Make sure it is a valid .xlsx generated by this system.")
        return redirect("import_attendance")

    if "Attendance" not in wb.sheetnames:
        messages.error(request, "Invalid template: 'Attendance' sheet not found.")
        return redirect("import_attendance")

    ws = wb["Attendance"]
    org = request.user.organization

    # Load org employees keyed by employee_id
    employees = {
        emp.employee_id: emp
        for emp in Employee.objects.filter(organization=org, is_active=True, is_deleted=False)
    }

    # Build date → column mapping from row 2 (merged date headers) and row 3 sub-headers
    # Row 2: merged cells with date label; row 3: Status / Clock In / Clock Out
    month_start = date(year, month, 1)
    month_end   = date(year, month, _cal_module.monthrange(year, month)[1])
    month_dates = {month_start + timedelta(days=i) for i in range((month_end - month_start).days + 1)}

    # Parse column layout from row 3 sub-headers
    # Columns 1-3: EmpID, Name, Dept; then triplets: Status, Clock In, Clock Out
    date_col_map = {}  # date -> (status_col, ci_col, co_col) all 1-indexed
    col = 4
    while col <= ws.max_column:
        # Find date from row 2 merged cell value
        cell_val = ws.cell(row=2, column=col).value
        if cell_val is None:
            col += 3
            continue
        # Extract the date portion (first line before \n)
        date_part = str(cell_val).split("\n")[0].strip()
        try:
            parsed_date = datetime.strptime(date_part, "%d %b").replace(year=year).date()
        except ValueError:
            col += 3
            continue
        if parsed_date in month_dates:
            date_col_map[parsed_date] = (col, col + 1, col + 2)
        col += 3

    if not date_col_map:
        messages.error(request, "Could not read date columns from the template. Make sure this file was generated by MULZON HRMS for the correct month/year.")
        return redirect("import_attendance")

    # Existing attendance for this month
    existing_att = {
        (att.employee_id, att.date): att
        for att in Attendance.objects.filter(
            employee__organization=org,
            date__range=(month_start, month_end),
        ).select_related("employee")
    }

    # Holidays
    holiday_dates = {
        h.date for h in Holiday.objects.filter(
            organization=org, date__range=(month_start, month_end), is_deleted=False
        )
    }

    ALLOWED_STATUSES = {"PRESENT", "ABSENT", "HALF DAY", "LEAVE", "HOLIDAY", "WEEKOFF", "N/A"}
    SKIP_STATUSES    = {"HOLIDAY", "WEEKOFF", "N/A", ""}

    records      = []
    errors       = []
    create_count = 0
    update_count = 0
    skip_count   = 0

    for row_num in range(4, ws.max_row + 1):
        emp_id_cell = ws.cell(row=row_num, column=1).value
        if emp_id_cell is None:
            break
        emp_id = str(emp_id_cell).strip()

        emp = employees.get(emp_id)
        if not emp:
            errors.append({
                "row": row_num,
                "emp_id": emp_id,
                "date": "—",
                "message": f"Employee ID '{emp_id}' not found in this organization.",
            })
            continue

        for d, (sc, cic, coc) in date_col_map.items():
            status_raw = ws.cell(row=row_num, column=sc).value
            ci_raw     = ws.cell(row=row_num, column=cic).value
            co_raw     = ws.cell(row=row_num, column=coc).value

            status = str(status_raw).strip().upper() if status_raw else ""

            if status in SKIP_STATUSES:
                skip_count += 1
                continue

            if status not in {s.upper() for s in ["Present", "Absent", "Half Day", "Leave"]}:
                errors.append({
                    "row": row_num,
                    "emp_id": emp_id,
                    "date": d.strftime("%d %b %Y"),
                    "message": f"Invalid status '{status_raw}'. Accepted: Present, Absent, Half Day, Leave.",
                })
                continue

            # Validate joining date
            if emp.joining_date and d < emp.joining_date:
                skip_count += 1
                continue

            clock_in  = _parse_excel_time(ci_raw)
            clock_out = _parse_excel_time(co_raw)

            # Absent: no times allowed
            if status == "ABSENT" and (clock_in or clock_out):
                errors.append({
                    "row": row_num,
                    "emp_id": emp_id,
                    "date": d.strftime("%d %b %Y"),
                    "message": "Absent cannot have Clock In or Clock Out times.",
                })
                continue

            # Present: clock_in required
            if status == "PRESENT" and not clock_in:
                errors.append({
                    "row": row_num,
                    "emp_id": emp_id,
                    "date": d.strftime("%d %b %Y"),
                    "message": "Present requires a Clock In time.",
                })
                continue

            # Clock out before clock in
            if clock_in and clock_out and clock_out <= clock_in:
                errors.append({
                    "row": row_num,
                    "emp_id": emp_id,
                    "date": d.strftime("%d %b %Y"),
                    "message": f"Clock Out ({clock_out}) cannot be earlier than Clock In ({clock_in}).",
                })
                continue

            existing = existing_att.get((emp.id, d))
            action   = "UPDATE" if existing else "CREATE"
            if action == "UPDATE":
                update_count += 1
            else:
                create_count += 1

            records.append({
                "row":        row_num,
                "emp_id":     emp_id,
                "emp_name":   f"{emp.first_name} {emp.last_name}",
                "date":       d.isoformat(),
                "date_label": d.strftime("%d %b %Y"),
                "status":     status,
                "clock_in":   clock_in.strftime("%H:%M") if clock_in else "",
                "clock_out":  clock_out.strftime("%H:%M") if clock_out else "",
                "action":     action,
                "att_id":     existing.id if existing else None,
            })

    # Store in session for confirmation step
    request.session["att_import_preview"] = {
        "month":        month,
        "year":         year,
        "records":      records,
        "errors":       errors,
        "create_count": create_count,
        "update_count": update_count,
        "skip_count":   skip_count,
        "file_name":    uploaded.name,
    }

    return redirect("import_attendance_preview")


@login_required
def import_attendance_preview_view(request):
    """Display the parsed preview data from session."""
    if not request.user.is_staff:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("Permission denied.")

    preview = request.session.get("att_import_preview")
    if not preview:
        messages.warning(request, "No import session found. Please upload a file first.")
        return redirect("import_attendance")

    return render(request, "attendance_import_preview.html", {
        "preview":      preview,
        "records":      preview.get("records", []),
        "errors":       preview.get("errors", []),
        "month_label":  date(preview["year"], preview["month"], 1).strftime("%B %Y"),
        "create_count": preview.get("create_count", 0),
        "update_count": preview.get("update_count", 0),
        "skip_count":   preview.get("skip_count", 0),
        "total":        len(preview.get("records", [])),
        "error_count":  len(preview.get("errors", [])),
    })


@login_required
def import_attendance_confirm_view(request):
    """Commit the previewed import atomically."""
    if not request.user.is_staff:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("Permission denied.")

    if request.method != "POST":
        return redirect("import_attendance")

    preview = request.session.get("att_import_preview")
    if not preview:
        messages.error(request, "Import session expired. Please upload again.")
        return redirect("import_attendance")

    conflict_strategy = request.POST.get("conflict_strategy", "skip")  # skip | update
    records           = preview.get("records", [])
    errors            = preview.get("errors", [])

    if errors and request.POST.get("abort_on_errors") == "1":
        messages.error(request, f"Import cancelled. Fix {len(errors)} error(s) and re-upload.")
        return redirect("import_attendance")

    org = request.user.organization

    # Reload employees + existing attendance fresh (session may be stale)
    employees = {
        emp.employee_id: emp
        for emp in Employee.objects.filter(organization=org, is_active=True, is_deleted=False)
    }

    if not records:
        messages.warning(request, "Nothing to import.")
        return redirect("import_attendance")

    month = preview["month"]
    year  = preview["year"]

    created_count = 0
    updated_count = 0
    skipped_count = 0

    from .models import AttendanceImportBatch, AttendanceAdminAction

    with transaction.atomic():
        batch = AttendanceImportBatch.objects.create(
            organization  = org,
            imported_by   = request.user,
            month         = month,
            year          = year,
            file_name     = preview.get("file_name", ""),
            total_records = len(records),
            status        = "VALIDATING",
            created_by    = request.user,
        )

        audit_bulk = []

        for rec in records:
            emp = employees.get(rec["emp_id"])
            if not emp:
                skipped_count += 1
                continue

            att_date  = date.fromisoformat(rec["date"])
            clock_in  = _parse_time(rec["clock_in"])  if rec["clock_in"]  else None
            clock_out = _parse_time(rec["clock_out"]) if rec["clock_out"] else None

            existing = Attendance.objects.filter(employee=emp, date=att_date).first()

            if existing:
                if conflict_strategy == "skip":
                    skipped_count += 1
                    continue
                # Update
                old_ci   = existing.clock_in
                old_co   = existing.clock_out
                old_late = existing.late_minutes

                existing.clock_in    = clock_in
                existing.clock_out   = clock_out
                existing.updated_by  = request.user
                existing.save()

                audit_bulk.append(AttendanceAdminAction(
                    organization     = org,
                    attendance       = existing,
                    employee         = emp,
                    performed_by     = request.user,
                    action_type      = "UPDATE",
                    attendance_date  = att_date,
                    old_clock_in     = old_ci,
                    old_clock_out    = old_co,
                    old_late_minutes = old_late,
                    new_clock_in     = clock_in,
                    new_clock_out    = clock_out,
                    new_late_minutes = 0,
                    created_by       = request.user,
                ))
                updated_count += 1

            else:
                att = Attendance(
                    organization = org,
                    employee     = emp,
                    date         = att_date,
                    clock_in     = clock_in,
                    clock_out    = clock_out,
                    late_minutes = 0,
                    created_by   = request.user,
                )
                att.save()

                audit_bulk.append(AttendanceAdminAction(
                    organization     = org,
                    attendance       = att,
                    employee         = emp,
                    performed_by     = request.user,
                    action_type      = "CREATE",
                    attendance_date  = att_date,
                    old_clock_in     = None,
                    old_clock_out    = None,
                    old_late_minutes = None,
                    new_clock_in     = clock_in,
                    new_clock_out    = clock_out,
                    new_late_minutes = 0,
                    created_by       = request.user,
                ))
                created_count += 1

        # Bulk create audit records
        AttendanceAdminAction.objects.bulk_create(audit_bulk, batch_size=500)

        # Update batch record
        batch.created_records = created_count
        batch.updated_records = updated_count
        batch.skipped_records = skipped_count
        batch.total_records   = created_count + updated_count + skipped_count
        batch.status          = "COMPLETED"
        batch.save()

    # Clear session
    del request.session["att_import_preview"]

    messages.success(
        request,
        f"Import complete: {created_count} created, {updated_count} updated, {skipped_count} skipped."
    )
    return redirect("import_attendance_history")


@login_required
def import_attendance_history_view(request):
    """Show all past import batches for this organization."""
    if not request.user.is_staff:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("Permission denied.")

    from .models import AttendanceImportBatch
    batches = AttendanceImportBatch.objects.filter(
        organization=request.user.organization
    ).select_related("imported_by").order_by("-created_at")[:50]

    return render(request, "attendance_import_history.html", {"batches": batches})


@login_required
def attendance_settings_view(request):
    """Attendance Settings – GET loads data, POST saves via AJAX."""
    if not request.user.is_staff:
        raise PermissionDenied("Only authorized staff can access this page.")

    from .models import AttendanceSettings
    import json

    att_settings, _ = AttendanceSettings.objects.get_or_create(
        organization=request.user.organization,
        defaults={
            'network_restriction_enabled': False,
            'allowed_ip_addresses': [],
            'location_restriction_enabled': False,
            'allowed_radius_meters': 100,
            'max_gps_accuracy_meters': 50,
        }
    )

    if request.method == 'POST':
        try:
            body = json.loads(request.body)
            errors = []

            # --- Network restriction ---
            att_settings.network_restriction_enabled = bool(body.get('network_restriction_enabled', False))
            raw_ips = body.get('allowed_ip_addresses', [])
            import ipaddress
            cleaned_ips = []
            for ip in raw_ips:
                ip = ip.strip()
                if not ip:
                    continue
                try:
                    ipaddress.ip_address(ip)
                    cleaned_ips.append(ip)
                except ValueError:
                    errors.append(f"'{ip}' is not a valid IP address.")
            att_settings.allowed_ip_addresses = cleaned_ips

            # --- Location restriction ---
            att_settings.location_restriction_enabled = bool(body.get('location_restriction_enabled', False))

            raw_lat = body.get('office_latitude', '')
            raw_lng = body.get('office_longitude', '')
            raw_radius = body.get('allowed_radius_meters', 100)
            raw_accuracy = body.get('max_gps_accuracy_meters', 50)

            # Validate coordinates only if location restriction is being enabled
            if att_settings.location_restriction_enabled:
                if raw_lat == '' or raw_lat is None:
                    errors.append("Office latitude is required when location restriction is enabled.")
                if raw_lng == '' or raw_lng is None:
                    errors.append("Office longitude is required when location restriction is enabled.")

            if raw_lat not in ('', None):
                try:
                    lat = float(raw_lat)
                    if not (-90 <= lat <= 90):
                        errors.append("Latitude must be between -90 and 90.")
                    else:
                        att_settings.office_latitude = lat
                except (ValueError, TypeError):
                    errors.append("Latitude must be a valid decimal number.")
            else:
                att_settings.office_latitude = None

            if raw_lng not in ('', None):
                try:
                    lng = float(raw_lng)
                    if not (-180 <= lng <= 180):
                        errors.append("Longitude must be between -180 and 180.")
                    else:
                        att_settings.office_longitude = lng
                except (ValueError, TypeError):
                    errors.append("Longitude must be a valid decimal number.")
            else:
                att_settings.office_longitude = None

            try:
                radius = int(raw_radius)
                if not (1 <= radius <= 50000):
                    errors.append("Allowed radius must be between 1 and 50,000 meters.")
                else:
                    att_settings.allowed_radius_meters = radius
            except (ValueError, TypeError):
                errors.append("Allowed radius must be a whole number.")

            try:
                accuracy = int(raw_accuracy)
                if not (5 <= accuracy <= 500):
                    errors.append("Maximum GPS accuracy must be between 5 and 500 meters.")
                else:
                    att_settings.max_gps_accuracy_meters = accuracy
            except (ValueError, TypeError):
                errors.append("Maximum GPS accuracy must be a whole number.")

            if errors:
                return JsonResponse({'success': False, 'errors': errors}, status=400)

            print("\n--- ATTENDANCE SETTINGS DEBUG ---")
            print("USER ORG ID:", getattr(request.user, 'organization_id', 'Unknown'))
            print("SETTINGS ID:", att_settings.id)
            print("INCOMING LAT/LNG:", raw_lat, raw_lng)
            print("INCOMING RADIUS:", raw_radius)
            print("INCOMING ACCURACY:", raw_accuracy)
            print("BEFORE SAVE RADIUS:", att_settings.allowed_radius_meters)
            print("BEFORE SAVE ACCURACY:", att_settings.max_gps_accuracy_meters)

            att_settings.save()
            att_settings.refresh_from_db()

            print("AFTER SAVE RADIUS:", att_settings.allowed_radius_meters)
            print("AFTER SAVE ACCURACY:", att_settings.max_gps_accuracy_meters)
            print("---------------------------------\n")

            return JsonResponse({'success': True, 'message': 'Settings saved successfully.'})
        except Exception as e:
            import logging
            logging.getLogger(__name__).error("attendance_settings_view POST error: %s", e)
            return JsonResponse({'success': False, 'errors': [str(e)]}, status=500)

    return render(request, 'attendance_settings.html', {'att_settings': att_settings})





