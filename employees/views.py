import csv
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.shortcuts import render , redirect
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.contrib import messages
from django.db.models import Q, Count
from .models import Department, Designation, Employee

# Create your views here.
@login_required
def employee_view(request):
    if not (request.user.is_staff or request.user.is_superuser):
        return redirect('employee_directory')
        
    # Base query
    employees_list = Employee.objects.filter(organization=request.user.organization, is_deleted=False).order_by('-created_at')
    
    # Get filter parameters
    search_query = request.GET.get('search', '')
    department_id = request.GET.get('department', '')
    designation_id = request.GET.get('designation', '')
    status = request.GET.get('status', '')
    
    # Apply filters
    if search_query:
        employees_list = employees_list.filter(
            Q(first_name__icontains=search_query) | 
            Q(last_name__icontains=search_query) | 
            Q(employee_id__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    
    if department_id:
        employees_list = employees_list.filter(department_id=department_id)
    
    if designation_id:
        employees_list = employees_list.filter(designation_id=designation_id)
        
    if status:
        employees_list = employees_list.filter(is_active=(status == 'Active'))

    # Stats for Dashboard
    stats = {
        'total': Employee.objects.filter(organization=request.user.organization, is_deleted=False).count(),
        'active': Employee.objects.filter(organization=request.user.organization, is_deleted=False, is_active=True).count(),
        'inactive': Employee.objects.filter(organization=request.user.organization, is_deleted=False, is_active=False).count(),
        'depts': Department.objects.filter(organization=request.user.organization, is_deleted=False).count()
    }

    # Pagination
    paginator = Paginator(employees_list, 10) # 10 employees per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Fetch context for filters
    departments = Department.objects.filter(organization=request.user.organization, is_deleted=False).order_by('name')
    designations = Designation.objects.filter(organization=request.user.organization, is_deleted=False).order_by('name')
    
    return render(request, 'employee.html', {
        'page_obj': page_obj,
        'stats': stats,
        'departments': departments,
        'designations': designations,
        'search_query': search_query,
        'selected_dept': department_id,
        'selected_desig': designation_id,
        'selected_status': status
    })

@login_required
def employee_directory(request):
    organization = request.user.organization
    
    # Show active, non-deleted employees of the organization (or global seeded ones)
    employees_list = Employee.objects.filter(
        Q(organization=organization) | Q(organization__isnull=True),
        is_active=True,
        is_deleted=False
    ).order_by('first_name', 'last_name')
    
    # Search filtering
    search_query = request.GET.get('search', '')
    if search_query:
        employees_list = employees_list.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(department__name__icontains=search_query) |
            Q(designation__name__icontains=search_query)
        )
        
    # Department filtering
    selected_dept = request.GET.get('department', '')
    if selected_dept:
        employees_list = employees_list.filter(department_id=selected_dept)
        
    # Designation filtering
    selected_desig = request.GET.get('designation', '')
    if selected_desig:
        employees_list = employees_list.filter(designation_id=selected_desig)
        
    # Paginator: 12 cards per page (3x4 grid)
    paginator = Paginator(employees_list, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Filter dropdown options
    departments = Department.objects.filter(
        Q(organization=organization) | Q(organization__isnull=True),
        is_deleted=False,
        is_active=True
    ).order_by('name')
    
    designations = Designation.objects.filter(
        Q(organization=organization) | Q(organization__isnull=True),
        is_deleted=False,
        is_active=True
    ).order_by('name')
    
    context = {
        'page_obj': page_obj,
        'departments': departments,
        'designations': designations,
        'search_query': search_query,
        'selected_dept': selected_dept,
        'selected_desig': selected_desig,
    }
    return render(request, 'employee_directory.html', context)

def department_view(request, pk=None):
    edit_dept = None
    if pk:
        edit_dept = Department.objects.get(pk=pk, organization=request.user.organization, is_deleted=False)

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        
        if not name:
            messages.error(request, "Department name is required.")
            return redirect('department')

        # Check across ALL records (including deleted) for uniqueness conflict
        # using iexact for case-insensitive protection
        existing_dept = Department.objects.filter(organization=request.user.organization, name__iexact=name).first()
        
        if existing_dept:
            if edit_dept and existing_dept.pk == edit_dept.pk:
                # Same record being edited, this is fine
                pass
            elif not existing_dept.is_deleted:
                # Conflict with another active department
                messages.error(request, f'An active department "{name}" already exists.')
                return redirect('department')
            else:
                # Conflict with a deleted record - Restore it instead of creating new
                existing_dept.is_deleted = False
                existing_dept.is_active = True
                existing_dept.description = description
                existing_dept.updated_by = request.user
                existing_dept.save()
                messages.success(request, f'Department "{name}" has been restored.')
                return redirect('department')

        if edit_dept:
            edit_dept.name = name
            edit_dept.description = description
            edit_dept.updated_by = request.user
            edit_dept.save()
            messages.success(request, f'Department "{name}" updated successfully!')
        else:
            Department.objects.create(
                organization=request.user.organization,
                name=name, 
                description=description,
                created_by=request.user
            )
            messages.success(request, f'Department "{name}" created successfully!')
        return redirect('department')
    
    departments = Department.objects.filter(organization=request.user.organization, is_deleted=False).order_by('-created_at')
    return render(request, 'department.html', {
        'departments': departments,
        'edit_dept': edit_dept
    })

def delete_department(request, pk):
    try:
        dept = Department.objects.get(pk=pk, organization=request.user.organization)
        name = dept.name
        dept.is_deleted = True
        dept.is_active = False
        dept.deleted_by = request.user
        dept.save()
        messages.success(request, f'Department "{name}" deleted successfully.')
    except Department.DoesNotExist:
        messages.error(request, 'Department not found.')
    return redirect('department')
def toggle_department_status(request, pk):
    try:
        dept = Department.objects.get(pk=pk, organization=request.user.organization)
        dept.is_active = not dept.is_active
        dept.updated_by = request.user
        dept.save()
        status = "activated" if dept.is_active else "deactivated"
        messages.success(request, f'Department "{dept.name}" {status} successfully.')
    except Department.DoesNotExist:
        messages.error(request, 'Department not found.')
    #if department have designations, then designations should also be deactivated
    designations = Designation.objects.filter(organization=request.user.organization, department=dept, is_deleted=False)
    for designation in designations:
        designation.is_active = False
        designation.updated_by = request.user
        designation.save()
    return redirect('department')

def designation_view(request, pk=None):
    edit_desig = None
    if pk:
        try:
            edit_desig = Designation.objects.get(pk=pk, organization=request.user.organization, is_deleted=False)
        except Designation.DoesNotExist:
            messages.error(request, 'Designation not found.')
            return redirect('designation')
        
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        dept_id = request.POST.get('department')
        if not dept_id:
            dept_id = None
        description = request.POST.get('description', '').strip()
        
        if not name:
            messages.error(request, "Designation name is required.")
            return redirect('designation')
        
        existing_desig = Designation.objects.filter(organization=request.user.organization, name__iexact=name).first()
        
        if existing_desig:
            if edit_desig and existing_desig.pk == edit_desig.pk:
                pass
            elif not existing_desig.is_deleted:
                messages.error(request, f'An active designation "{name}" already exists.')
                return redirect('designation')
            else:
                existing_desig.is_deleted = False
                existing_desig.is_active = True
                existing_desig.description = description
                existing_desig.department_id = dept_id
                existing_desig.updated_by = request.user
                existing_desig.save()
                messages.success(request, f'Designation "{name}" has been restored.')
                return redirect('designation')

        if edit_desig:
            edit_desig.name = name
            edit_desig.description = description
            edit_desig.department_id = dept_id
            edit_desig.updated_by = request.user
            edit_desig.save()
            messages.success(request, f'Designation "{name}" updated successfully!')
        else:
            Designation.objects.create(
                organization=request.user.organization,
                name=name, 
                description=description,
                department_id=dept_id,
                created_by=request.user
            )
            messages.success(request, f'Designation "{name}" created successfully!')
        return redirect('designation')

    designations = Designation.objects.filter(organization=request.user.organization, is_deleted=False).order_by('-created_at')
    departments = Department.objects.filter(organization=request.user.organization, is_deleted=False).order_by('name')
    
    return render(request, 'designation.html', {
        'designations': designations,
        'departments': departments,
        'edit_desig': edit_desig
    })
def delete_designation(request, pk):
    try:
        desig = Designation.objects.get(pk=pk, organization=request.user.organization)
        name = desig.name
        desig.is_deleted = True
        desig.is_active = False
        desig.deleted_by = request.user
        desig.save()
        messages.success(request, f'Designation "{name}" deleted successfully.')
    except Designation.DoesNotExist:
        messages.error(request, 'Designation not found.')
    return redirect('designation')
def toggle_designation_status(request, pk):
    try:
        desig = Designation.objects.get(pk=pk, organization=request.user.organization)
        desig.is_active = not desig.is_active
        desig.updated_by = request.user
        desig.save()
        status = "activated" if desig.is_active else "deactivated"
        messages.success(request, f'Designation "{desig.name}" {status} successfully.')
    except Designation.DoesNotExist:
        messages.error(request, 'Designation not found.')
    #if designation is active and department is inactive then department should be activated
    if desig.is_active and not desig.department.is_active:
        desig.department.is_active = True
        desig.department.updated_by = request.user
        desig.department.save()
    return redirect('designation')
def add_employee(request):
    if request.method == 'POST':
        try:
            # Instantiate without saving yet to catch errors in a controlled way
            employee = Employee(
                organization=request.user.organization,
                employee_id=request.POST.get('employee_id'),
                first_name=request.POST.get('first_name'),
                last_name=request.POST.get('last_name'),
                email=request.POST.get('email'),
                phone_number=request.POST.get('phone_number'),
                date_of_birth=request.POST.get('date_of_birth') or None,
                gender=request.POST.get('gender'),
                marital_status=request.POST.get('marital_status'),
                blood_group=request.POST.get('blood_group'),
                nationality=request.POST.get('nationality'),
                current_address=request.POST.get('current_address'),
                current_city=request.POST.get('current_city'),
                current_state=request.POST.get('current_state'),
                current_country=request.POST.get('current_country'),
                current_pincode=request.POST.get('current_pincode'),
                permanent_address=request.POST.get('permanent_address'),
                permanent_city=request.POST.get('permanent_city'),
                permanent_state=request.POST.get('permanent_state'),
                permanent_country=request.POST.get('permanent_country'),
                permanent_pincode=request.POST.get('permanent_pincode'),
                joining_date=request.POST.get('joining_date') or None,
                employment_type=request.POST.get('employment_type'),
                department_id=request.POST.get('department') or None,
                designation_id=request.POST.get('designation') or None,
                manager_id=request.POST.get('manager') or None,
                work_location=request.POST.get('work_location'),
                pan_number=request.POST.get('pan_number'),
                aadhaar_number=request.POST.get('aadhaar_number'),
                bank_name=request.POST.get('bank_name'),
                bank_account_number=request.POST.get('bank_account_number'),
                ifsc_code=request.POST.get('ifsc_code'),
                emergency_contact_name=request.POST.get('emergency_contact_name'),
                emergency_contact_number=request.POST.get('emergency_contact_number'),
                emergency_contact_relationship=request.POST.get('emergency_contact_relationship'),
                resume=request.FILES.get('resume'),
                offer_letter=request.FILES.get('offer_letter'),
                aadhaar_card=request.FILES.get('aadhaar_card'),
                pan_card=request.FILES.get('pan_card'),
                appointment_letter=request.FILES.get('appointment_letter'),
                profile_img=request.FILES.get('profile_img'),
                created_by=request.user
            )
            employee.save()
            messages.success(request, f'Employee "{employee.first_name} {employee.last_name}" created successfully!')
            return redirect('employee')
        except Exception as e:
            messages.error(request, f'Failed to create employee: {str(e)}')
            # Re-fetch context for the template
            departments = Department.objects.filter(organization=request.user.organization, is_deleted=False).order_by('name')
            designations = Designation.objects.filter(organization=request.user.organization, is_deleted=False).order_by('name')
            managers = Employee.objects.filter(organization=request.user.organization, is_deleted=False).order_by('first_name')
            
            # Render the same page with current POST data and error messages
            return render(request, 'addemployee.html', {
                'departments': departments,
                'designations': designations,
                'managers': managers,
                'form_data': request.POST # Pass back data to preserve it
            })

    departments = Department.objects.filter(organization=request.user.organization, is_deleted=False).order_by('name')
    designations = Designation.objects.filter(organization=request.user.organization, is_deleted=False).order_by('name')
    managers = Employee.objects.filter(organization=request.user.organization, is_deleted=False).order_by('first_name')
    
    return render(request, 'addemployee.html', {
        'departments': departments,
        'designations': designations,
        'managers': managers
    })

@login_required
def toggle_employee_status(request, pk):
    try:
        employee = Employee.objects.get(pk=pk, organization=request.user.organization, is_deleted=False)
        employee.is_active = not employee.is_active
        employee.save()
        status_str = "activated" if employee.is_active else "deactivated"
        messages.success(request, f'Employee "{employee.first_name}" {status_str} successfully.')
    except Employee.DoesNotExist:
        messages.error(request, 'Employee not found.')
    return redirect('employee')

@login_required
def delete_employee(request, pk):
    try:
        employee = Employee.objects.get(pk=pk, organization=request.user.organization)
        employee.is_deleted = True
        employee.deleted_by = request.user
        employee.save()
        messages.success(request, f'Employee "{employee.first_name} {employee.last_name}" deleted successfully.')
    except Employee.DoesNotExist:
        messages.error(request, 'Employee not found.')
    return redirect('employee')

@login_required
def view_employee(request, pk):
    try:
        employee = Employee.objects.get(pk=pk, organization=request.user.organization, is_deleted=False)
    except Employee.DoesNotExist:
        messages.error(request, 'Employee not found.')
        return redirect('employee')
    
    return render(request, 'viewemployee.html', {
        'employee': employee
    })

@login_required
def edit_employee(request, pk):
    try:
        employee = Employee.objects.get(pk=pk, organization=request.user.organization, is_deleted=False)
    except Employee.DoesNotExist:
        messages.error(request, 'Employee not found.')
        return redirect('employee')

    if request.method == 'POST':
        try:
            employee.employee_id = request.POST.get('employee_id')
            employee.first_name = request.POST.get('first_name')
            employee.last_name = request.POST.get('last_name')
            employee.email = request.POST.get('email')
            employee.phone_number = request.POST.get('phone_number')
            employee.date_of_birth = request.POST.get('date_of_birth') or None
            employee.gender = request.POST.get('gender')
            employee.marital_status = request.POST.get('marital_status')
            employee.blood_group = request.POST.get('blood_group')
            employee.nationality = request.POST.get('nationality')
            employee.current_address = request.POST.get('current_address')
            employee.current_city = request.POST.get('current_city')
            employee.current_state = request.POST.get('current_state')
            employee.current_country = request.POST.get('current_country')
            employee.current_pincode = request.POST.get('current_pincode')
            employee.permanent_address = request.POST.get('permanent_address')
            employee.permanent_city = request.POST.get('permanent_city')
            employee.permanent_state = request.POST.get('permanent_state')
            employee.permanent_country = request.POST.get('permanent_country')
            employee.permanent_pincode = request.POST.get('permanent_pincode')
            employee.joining_date = request.POST.get('joining_date') or None
            employee.employment_type = request.POST.get('employment_type')
            employee.department_id = request.POST.get('department') or None
            employee.designation_id = request.POST.get('designation') or None
            employee.manager_id = request.POST.get('manager') or None
            employee.work_location = request.POST.get('work_location')
            employee.pan_number = request.POST.get('pan_number')
            employee.aadhaar_number = request.POST.get('aadhaar_number')
            employee.bank_name = request.POST.get('bank_name')
            employee.bank_account_number = request.POST.get('bank_account_number')
            employee.ifsc_code = request.POST.get('ifsc_code')
            employee.emergency_contact_name = request.POST.get('emergency_contact_name')
            employee.emergency_contact_number = request.POST.get('emergency_contact_number')
            employee.emergency_contact_relationship = request.POST.get('emergency_contact_relationship')
            
            if request.FILES.get('resume'):
                employee.resume = request.FILES.get('resume')
            if request.FILES.get('offer_letter'):
                employee.offer_letter = request.FILES.get('offer_letter')
            if request.FILES.get('aadhaar_card'):
                employee.aadhaar_card = request.FILES.get('aadhaar_card')
            if request.FILES.get('pan_card'):
                employee.pan_card = request.FILES.get('pan_card')
            if request.FILES.get('appointment_letter'):
                employee.appointment_letter = request.FILES.get('appointment_letter')
            if request.FILES.get('profile_img'):
                employee.profile_img = request.FILES.get('profile_img')

            employee.updated_by = request.user
            employee.save()
            messages.success(request, f'Employee "{employee.first_name} {employee.last_name}" updated successfully!')
            return redirect('view_employee', pk=employee.pk)
        except Exception as e:
            messages.error(request, f'Failed to update employee: {str(e)}')

    departments = Department.objects.filter(organization=request.user.organization, is_deleted=False).order_by('name')
    designations = Designation.objects.filter(organization=request.user.organization, is_deleted=False).order_by('name')
    managers = Employee.objects.filter(organization=request.user.organization, is_deleted=False).exclude(pk=employee.pk).order_by('first_name')
    
    return render(request, 'editemployee.html', {
        'employee': employee,
        'departments': departments,
        'designations': designations,
        'managers': managers
    })

@login_required
def export_employees_csv(request):
    # Base query
    employees = Employee.objects.filter(organization=request.user.organization, is_deleted=False).order_by('-created_at')
    
    # Get filter parameters
    search_query = request.GET.get('search', '')
    department_id = request.GET.get('department', '')
    designation_id = request.GET.get('designation', '')
    status = request.GET.get('status', '')
    
    # Apply filters
    if search_query:
        employees = employees.filter(
            Q(first_name__icontains=search_query) | 
            Q(last_name__icontains=search_query) | 
            Q(employee_id__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    
    if department_id:
        employees = employees.filter(department_id=department_id)
    
    if designation_id:
        employees = employees.filter(designation_id=designation_id)
        
    if status:
        is_active = (status == 'Active')
        employees = employees.filter(is_active=is_active)

    # Create CSV
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="employees_export.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Employee ID', 'First Name', 'Last Name', 'Email', 'Phone', 
        'Department', 'Designation', 'Joining Date', 'Employment Type', 'Status'
    ])
    
    for emp in employees:
        writer.writerow([
            emp.employee_id,
            emp.first_name,
            emp.last_name,
            emp.email,
            emp.phone_number,
            emp.department.name if emp.department else '--',
            emp.designation.name if emp.designation else '--',
            emp.joining_date,
            emp.employment_type,
            'Active' if emp.is_active else 'Inactive'
        ])
        
    return response

@login_required
def export_employees_excel(request):
    # Base query
    employees = Employee.objects.filter(organization=request.user.organization, is_deleted=False).order_by('-created_at')
    
    # Same filtering logic
    search_query = request.GET.get('search', '')
    department_id = request.GET.get('department', '')
    designation_id = request.GET.get('designation', '')
    status = request.GET.get('status', '')
    
    if search_query:
        employees = employees.filter(
            Q(first_name__icontains=search_query) | 
            Q(last_name__icontains=search_query) | 
            Q(employee_id__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    if department_id:
        employees = employees.filter(department_id=department_id)
    if designation_id:
        employees = employees.filter(designation_id=designation_id)
    if status:
        employees = employees.filter(is_active=(status == 'Active'))

    # Create Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Directory"

    # Define Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    center_aligned = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin', color="E2E8F0"), 
        right=Side(style='thin', color="E2E8F0"), 
        top=Side(style='thin', color="E2E8F0"), 
        bottom=Side(style='thin', color="E2E8F0")
    )

    # Header Row
    headers = ['Employee ID', 'Full Name', 'Email Address', 'Phone Number', 'Department', 'Designation', 'Joining Date', 'Type', 'Status']
    ws.append(headers)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_aligned
        cell.border = border

    # Data Rows
    for index, emp in enumerate(employees, start=2):
        row_data = [
            emp.employee_id,
            f"{emp.first_name} {emp.last_name}",
            emp.email,
            emp.phone_number,
            emp.department.name if emp.department else '--',
            emp.designation.name if emp.designation else '--',
            str(emp.joining_date) if emp.joining_date else '--',
            emp.employment_type,
            'Active' if emp.is_active else 'Inactive'
        ]
        ws.append(row_data)

        # Style data row
        fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") if index % 2 == 0 else None
        for cell in ws[index]:
            cell.border = border
            if fill:
                cell.fill = fill
            
            # Highlight status
            if cell.column == 9: # Status column
                if cell.value == 'Active':
                    cell.font = Font(bold=True, color="059669") # Green
                else:
                    cell.font = Font(bold=True, color="DC2626") # Red

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 4)
        ws.column_dimensions[column].width = adjusted_width

    # Prepare Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Employee_Directory_Export.xlsx"'
    wb.save(response)
    
    return response

@login_required
def download_sample_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bulk Import Template"

    # Define Headers
    headers = [
        'Employee ID*', 'First Name*', 'Last Name*', 'Email*', 'Phone*', 
        'Department', 'Designation', 'Joining Date (YYYY-MM-DD)', 
        'Employment Type', 'Gender', 'Work Location'
    ]
    ws.append(headers)

    # Style Headers
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Add Sample Row
    sample_row = [
        'EMP1001', 'John', 'Doe', 'john.doe@example.com', '1234567890',
        'IT', 'Software Engineer', '2024-01-01', 'Full Time', 'Male', 'Mumbai'
    ]
    ws.append(sample_row)

    # Auto-adjust column widths
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Employee_Import_Template.xlsx"'
    wb.save(response)
    return response

@login_required
def bulk_import(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        
        if not excel_file:
            messages.error(request, "Please upload an Excel file.")
            return redirect('bulk_import')
            
        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, "Format not supported. Please upload a .xlsx file.")
            return redirect('bulk_import')

        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_file)
            ws = wb.active
            
            imported_count = 0
            error_count = 0
            errors = []

            # Skip header row
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row): continue # Skip empty rows

                emp_id, f_name, l_name, email, phone, dept_name, desig_name, join_date, emp_type, gender, loc = row[:11]

                # Basic validation
                if not all([emp_id, f_name, l_name, email]):
                    errors.append(f"Row {row_idx}: Missing required fields.")
                    error_count += 1
                    continue

                if Employee.objects.filter(organization=request.user.organization, employee_id=emp_id).exists():
                    errors.append(f"Row {row_idx}: Employee ID '{emp_id}' already exists.")
                    error_count += 1
                    continue
                
                if Employee.objects.filter(organization=request.user.organization, email=email).exists():
                    errors.append(f"Row {row_idx}: Email '{email}' already exists.")
                    error_count += 1
                    continue

                # Process Foreign Keys
                dept = None
                if dept_name:
                    dept, _ = Department.objects.get_or_create(organization=request.user.organization, name=dept_name, defaults={'created_by': request.user})
                
                desig = None
                if desig_name:
                    desig, _ = Designation.objects.get_or_create(organization=request.user.organization, name=desig_name, defaults={'department': dept, 'created_by': request.user})

                try:
                    Employee.objects.create(
                        organization=request.user.organization,
                        employee_id=emp_id,
                        first_name=f_name,
                        last_name=l_name,
                        email=email,
                        phone_number=str(phone),
                        department=dept,
                        designation=desig,
                        joining_date=join_date if join_date else None,
                        employment_type=emp_type,
                        gender=gender,
                        work_location=loc,
                        created_by=request.user
                    )
                    imported_count += 1
                except Exception as e:
                    errors.append(f"Row {row_idx}: {str(e)}")
                    error_count += 1

            if imported_count > 0:
                messages.success(request, f"Successfully imported {imported_count} employees.")
            if error_count > 0:
                messages.warning(request, f"Skipped {error_count} rows due to errors.")
                for err in errors[:5]: # Show first 5 errors
                    messages.error(request, err)
            
            return redirect('employee')

        except Exception as e:
            messages.error(request, f"Error processing file: {str(e)}")
            return redirect('bulk_import')

    return render(request, 'bulk_import.html')

@login_required
def approve_employee(request, pk):
    from accounts.models import CustomUser
    from attendance.models import Shift, ShiftAssignment
    
    if not request.user.is_staff:
        return redirect('home')
        
    try:
        pending_user = CustomUser.objects.get(pk=pk, is_approved=False)
    except CustomUser.DoesNotExist:
        messages.error(request, 'Pending user request not found.')
        return redirect('home')
        
    if request.method == 'POST':
        dept_id = request.POST.get('department')
        desig_id = request.POST.get('designation')
        shift_id = request.POST.get('shift')
        
        if not dept_id or not desig_id or not shift_id:
            messages.error(request, 'Department, Designation, and Shift are required.')
            return redirect('approve_employee', pk=pk)
            
        # Create Employee
        import uuid
        emp = Employee.objects.create(
            employee_id=f"EMP-{str(uuid.uuid4())[:4].upper()}",
            first_name=pending_user.username,
            email=pending_user.email,
            phone_number=pending_user.phone or "",
            department_id=dept_id,
            designation_id=desig_id,
            organization=pending_user.organization,
            created_by=request.user
        )
        
        # Assign Shift
        from datetime import datetime
        shift = Shift.objects.get(pk=shift_id)
        ShiftAssignment.objects.create(
            employee=emp,
            shift=shift,
            effective_from=datetime.now().date(),
            created_by=request.user
        )
        
        # Approve User
        pending_user.is_approved = True
        pending_user.save()
        
        messages.success(request, f'Employee {pending_user.username} approved successfully!')
        return redirect('home')
        
    # GET request
    departments = Department.objects.filter(organization=request.user.organization, is_deleted=False)
    designations = Designation.objects.filter(organization=request.user.organization, is_deleted=False)
    shifts = Shift.objects.filter(organization=request.user.organization, is_deleted=False)
    
    return render(request, 'approve_employee.html', {
        'pending_user': pending_user,
        'departments': departments,
        'designations': designations,
        'shifts': shifts
    })

@login_required
def reject_employee(request, pk):
    from accounts.models import CustomUser
    if not request.user.is_staff:
        return redirect('home')
        
    try:
        pending_user = CustomUser.objects.get(pk=pk, is_approved=False)
        pending_user.is_active = False
        pending_user.save()
        messages.success(request, 'Employee request rejected and deactivated.')
    except CustomUser.DoesNotExist:
        messages.error(request, 'Pending user request not found.')
        
    return redirect('home')
@login_required
def edit_profile(request):
    organization = request.user.organization
    employee = Employee.objects.filter(email=request.user.email, organization=organization, is_active=True, is_deleted=False).first()
    
    if not employee:
        messages.error(request, "Employee record not found for your account.")
        return redirect('home')

    if request.method == 'POST':
        try:
            # Basic Info
            employee.first_name = request.POST.get('first_name')
            employee.last_name = request.POST.get('last_name')
            employee.phone_number = request.POST.get('phone_number')
            employee.date_of_birth = request.POST.get('date_of_birth') or None
            employee.gender = request.POST.get('gender')
            employee.marital_status = request.POST.get('marital_status')
            employee.blood_group = request.POST.get('blood_group')
            employee.nationality = request.POST.get('nationality')
            
            # Address
            employee.current_address = request.POST.get('current_address')
            employee.current_city = request.POST.get('current_city')
            employee.current_state = request.POST.get('current_state')
            employee.current_pincode = request.POST.get('current_pincode')
            employee.permanent_address = request.POST.get('permanent_address')
            employee.permanent_city = request.POST.get('permanent_city')
            employee.permanent_state = request.POST.get('permanent_state')
            employee.permanent_pincode = request.POST.get('permanent_pincode')
            
            # Financial
            employee.pan_number = request.POST.get('pan_number')
            employee.aadhaar_number = request.POST.get('aadhaar_number')
            employee.bank_name = request.POST.get('bank_name')
            employee.bank_account_number = request.POST.get('bank_account_number')
            employee.ifsc_code = request.POST.get('ifsc_code')
            
            # Emergency
            employee.emergency_contact_name = request.POST.get('emergency_contact_name')
            employee.emergency_contact_number = request.POST.get('emergency_contact_number')
            employee.emergency_contact_relationship = request.POST.get('emergency_contact_relationship')

            # Files
            if request.POST.get('remove_profile_img') == '1':
                employee.profile_img = None
            elif request.FILES.get('profile_img'):
                employee.profile_img = request.FILES.get('profile_img')
                
            if request.FILES.get('resume'):
                employee.resume = request.FILES.get('resume')
            if request.FILES.get('aadhaar_card'):
                employee.aadhaar_card = request.FILES.get('aadhaar_card')
            if request.FILES.get('pan_card'):
                employee.pan_card = request.FILES.get('pan_card')

            employee.updated_by = request.user
            employee.save()
            messages.success(request, "Your profile has been updated successfully!")
            return redirect('home')
        except Exception as e:
            messages.error(request, f"Failed to update profile: {str(e)}")

    return render(request, 'edit_profile.html', {
        'employee': employee
    })
