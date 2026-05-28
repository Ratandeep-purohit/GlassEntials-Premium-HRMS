from django.db import transaction
from django.utils.dateparse import parse_date
from django.utils.datastructures import MultiValueDict

from employees.models import Department, Designation
from leaves.models import (
    LeaveAccrualRule,
    LeaveCarryForwardRule,
    LeaveCategory,
    LeaveDurationRule,
    LeaveEligibilityRule,
    LeaveIntegrationRule,
    LeavePolicy,
    LeaveProofRule,
    LeaveRestrictionRule,
    LeaveSandwichRule,
    LeaveType,
    LeaveWorkflow,
    LeaveWorkflowStep,
)
from .audit_service import LeaveAuditService
from .utils import as_decimal, as_int, checkbox


class LeavePolicyWorkbenchService:
    EXCEL_IMPORT_COLUMNS = [
        "name",
        "code",
        "category_name",
        "category_code",
        "status",
        "description",
        "color_tag",
        "financial_year_start_month",
        "effective_from",
        "effective_to",
        "policy_version",
        "max_balance",
        "negative_balance_limit",
        "is_paid",
        "is_statutory",
        "is_requestable",
        "allow_negative_balance",
        "accrual_frequency",
        "accrual_enabled",
        "accrual_rate",
        "accrual_day",
        "max_yearly_accrual",
        "max_balance_cap",
        "prorate_on_joining",
        "prorate_on_exit",
        "rounding_mode",
        "eligible_gender",
        "min_service_days",
        "probation_allowed",
        "confirmation_required",
        "employee_filter_mode",
        "employment_types",
        "work_locations",
        "min_duration",
        "max_duration",
        "max_consecutive_days",
        "backdate_allowed",
        "max_backdate_days",
        "future_date_allowed",
        "max_future_days",
        "block_during_payroll_lock",
        "carry_forward_allowed",
        "carry_forward_limit",
        "max_carry_forward_days",
        "carry_expiry_month",
        "carry_expiry_day",
        "lapse_remaining",
        "proof_required",
        "required_after_days",
        "allowed_file_types",
        "max_file_size_mb",
        "approval_route",
        "approval_mode",
        "approval_sla_hours",
        "delegation_allowed",
        "attendance_sync_enabled",
        "prevent_attendance_conflict",
        "adjust_late_marks",
        "payroll_sync_enabled",
        "generate_lop",
        "negative_balance_deduction",
        "enable_encashment",
    ]

    EXCEL_REQUIRED_COLUMNS = {"name", "code", "max_balance"}
    BOOLEAN_COLUMNS = {
        "is_paid",
        "is_statutory",
        "is_requestable",
        "allow_negative_balance",
        "accrual_enabled",
        "prorate_on_joining",
        "prorate_on_exit",
        "probation_allowed",
        "confirmation_required",
        "backdate_allowed",
        "future_date_allowed",
        "block_during_payroll_lock",
        "carry_forward_allowed",
        "lapse_remaining",
        "proof_required",
        "delegation_allowed",
        "attendance_sync_enabled",
        "prevent_attendance_conflict",
        "adjust_late_marks",
        "payroll_sync_enabled",
        "generate_lop",
        "negative_balance_deduction",
        "enable_encashment",
    }
    LIST_COLUMNS = {"employment_types"}
    DEFAULT_IMPORT_VALUES = {
        "category_name": "Paid Leave",
        "category_code": "PAID",
        "status": "ACTIVE",
        "color_tag": "#2563eb",
        "financial_year_start_month": "4",
        "policy_version": "1",
        "negative_balance_limit": "0",
        "is_paid": "on",
        "is_requestable": "on",
        "accrual_frequency": "MONTHLY",
        "accrual_enabled": "on",
        "accrual_rate": "0",
        "accrual_day": "1",
        "max_yearly_accrual": "0",
        "max_balance_cap": "0",
        "prorate_on_joining": "on",
        "prorate_on_exit": "on",
        "rounding_mode": "NONE",
        "eligible_gender": "ANY",
        "min_service_days": "0",
        "probation_allowed": "on",
        "employee_filter_mode": "ALL",
        "min_duration": "0.5",
        "max_duration": "0",
        "max_consecutive_days": "0",
        "backdate_allowed": "on",
        "max_backdate_days": "0",
        "future_date_allowed": "on",
        "max_future_days": "365",
        "block_during_payroll_lock": "on",
        "carry_forward_limit": "0",
        "max_carry_forward_days": "0",
        "carry_expiry_month": "3",
        "carry_expiry_day": "31",
        "lapse_remaining": "on",
        "required_after_days": "0",
        "allowed_file_types": "pdf,jpg,jpeg,png",
        "max_file_size_mb": "5",
        "approval_route": "MANAGER_HR",
        "approval_mode": "ANY_ONE",
        "approval_sla_hours": "24",
        "delegation_allowed": "on",
        "attendance_sync_enabled": "on",
        "prevent_attendance_conflict": "on",
        "adjust_late_marks": "on",
        "payroll_sync_enabled": "on",
        "generate_lop": "on",
    }

    @staticmethod
    def save_from_post(*, organization, user, data, request=None):
        name = (data.get("name") or "").strip()
        code = (data.get("code") or "").strip().upper()
        if not name:
            raise ValueError("Leave name is required.")
        if not code:
            raise ValueError("Leave code is required.")

        category_name = (data.get("category_name") or "General Leave").strip()
        category_code = (data.get("category_code") or category_name[:10] or "GEN").strip().upper().replace(" ", "_")
        category, _ = LeaveCategory.objects.update_or_create(
            organization=organization,
            code=category_code,
            defaults={
                "name": category_name,
                "description": data.get("category_description", ""),
                "color": data.get("color_tag") or "#2563eb",
                "is_paid_category": checkbox(data, "is_paid"),
                "updated_by": user,
                "created_by": user,
            },
        )

        leave_type, created = LeaveType.objects.update_or_create(
            organization=organization,
            code=code,
            policy_version=as_int(data.get("policy_version"), 1),
            defaults={
                "category": category,
                "name": name,
                "description": data.get("description", ""),
                "color_tag": data.get("color_tag") or "#2563eb",
                "financial_year_start_month": as_int(data.get("financial_year_start_month"), 4),
                "effective_from": parse_date(data.get("effective_from") or "") or None,
                "effective_to": parse_date(data.get("effective_to") or "") or None,
                "status": data.get("status") or "DRAFT",
                "is_paid": checkbox(data, "is_paid"),
                "is_statutory": checkbox(data, "is_statutory"),
                "is_requestable": checkbox(data, "is_requestable"),
                "allow_negative_balance": checkbox(data, "allow_negative_balance"),
                "negative_balance_limit": as_decimal(data.get("negative_balance_limit")),
                "updated_by": user,
                "created_by": user,
            },
        )

        LeavePolicy.objects.update_or_create(
            organization=organization,
            leave_type=leave_type,
            defaults={
                "accrual_rate": as_decimal(data.get("accrual_rate")),
                "max_balance": as_decimal(data.get("max_balance")),
                "carry_forward_limit": as_decimal(data.get("carry_forward_limit")),
                "min_service_days": as_int(data.get("min_service_days")),
                "sandwich_rule": checkbox(data, "count_weekends_between_leave") or checkbox(data, "count_holidays_between_leave"),
                "requires_attachment": checkbox(data, "proof_required"),
                "attachment_threshold_days": as_int(data.get("required_after_days"), 3),
                "updated_by": user,
                "created_by": user,
            },
        )

        LeaveAccrualRule.objects.update_or_create(
            organization=organization,
            leave_type=leave_type,
            defaults={
                "frequency": data.get("accrual_frequency") or "MONTHLY",
                "accrual_day": as_int(data.get("accrual_day"), 1),
                "accrual_month": as_int(data.get("accrual_month"), 0) or None,
                "accrual_rate": as_decimal(data.get("accrual_rate")),
                "max_yearly_accrual": as_decimal(data.get("max_yearly_accrual")),
                "max_balance_cap": as_decimal(data.get("max_balance_cap") or data.get("max_balance")),
                "prorate_on_joining": checkbox(data, "prorate_on_joining"),
                "prorate_on_exit": checkbox(data, "prorate_on_exit"),
                "rounding_mode": data.get("rounding_mode") or "NONE",
                "enabled": checkbox(data, "accrual_enabled"),
                "updated_by": user,
                "created_by": user,
            },
        )

        LeaveCarryForwardRule.objects.update_or_create(
            organization=organization,
            leave_type=leave_type,
            defaults={
                "is_allowed": checkbox(data, "carry_forward_allowed"),
                "max_carry_forward_days": as_decimal(data.get("max_carry_forward_days") or data.get("carry_forward_limit")),
                "expiry_month": as_int(data.get("carry_expiry_month"), 3),
                "expiry_day": as_int(data.get("carry_expiry_day"), 31),
                "encash_remaining": checkbox(data, "encash_remaining"),
                "lapse_remaining": checkbox(data, "lapse_remaining"),
                "enabled": checkbox(data, "carry_forward_enabled"),
                "updated_by": user,
                "created_by": user,
            },
        )

        eligibility, _ = LeaveEligibilityRule.objects.update_or_create(
            organization=organization,
            leave_type=leave_type,
            defaults={
                "gender": data.get("eligible_gender") or "ANY",
                "employment_types": data.getlist("employment_types") if hasattr(data, "getlist") else [],
                "work_locations": [item.strip() for item in (data.get("work_locations") or "").split(",") if item.strip()],
                "min_service_days": as_int(data.get("min_service_days")),
                "probation_allowed": checkbox(data, "probation_allowed"),
                "confirmation_required": checkbox(data, "confirmation_required"),
                "employee_filter_mode": data.get("employee_filter_mode") or "ALL",
                "enabled": checkbox(data, "eligibility_enabled"),
                "updated_by": user,
                "created_by": user,
            },
        )
        eligibility.departments.set(Department.objects.filter(organization=organization, id__in=data.getlist("department_ids") if hasattr(data, "getlist") else []))
        eligibility.designations.set(Designation.objects.filter(organization=organization, id__in=data.getlist("designation_ids") if hasattr(data, "getlist") else []))

        LeaveRestrictionRule.objects.update_or_create(
            organization=organization,
            leave_type=leave_type,
            defaults={
                "min_duration": as_decimal(data.get("min_duration"), "0.50"),
                "max_duration": as_decimal(data.get("max_duration")),
                "max_consecutive_days": as_decimal(data.get("max_consecutive_days")),
                "min_gap_between_requests": as_int(data.get("min_gap_between_requests")),
                "max_requests_per_month": as_int(data.get("max_requests_per_month")),
                "max_requests_per_year": as_int(data.get("max_requests_per_year")),
                "backdate_allowed": checkbox(data, "backdate_allowed"),
                "max_backdate_days": as_int(data.get("max_backdate_days")),
                "future_date_allowed": checkbox(data, "future_date_allowed"),
                "max_future_days": as_int(data.get("max_future_days"), 365),
                "block_during_payroll_lock": checkbox(data, "block_during_payroll_lock"),
                "block_month_end": checkbox(data, "block_month_end"),
                "blackout_periods": [item.strip() for item in (data.get("blackout_periods") or "").splitlines() if item.strip()],
                "enabled": checkbox(data, "restriction_enabled"),
                "updated_by": user,
                "created_by": user,
            },
        )

        LeaveDurationRule.objects.update_or_create(
            organization=organization,
            leave_type=leave_type,
            defaults={
                "allow_full_day": checkbox(data, "allow_full_day"),
                "allow_half_day": checkbox(data, "allow_half_day"),
                "allow_hourly": checkbox(data, "allow_hourly"),
                "minimum_hourly_unit_minutes": as_int(data.get("minimum_hourly_unit_minutes"), 60),
                "maximum_hourly_duration_minutes": as_int(data.get("maximum_hourly_duration_minutes"), 480),
                "shift_aware": checkbox(data, "shift_aware"),
                "allowed_sessions": data.getlist("allowed_sessions") if hasattr(data, "getlist") else [],
                "updated_by": user,
                "created_by": user,
            },
        )

        LeaveSandwichRule.objects.update_or_create(
            organization=organization,
            leave_type=leave_type,
            defaults={
                "count_weekends_between_leave": checkbox(data, "count_weekends_between_leave"),
                "count_holidays_between_leave": checkbox(data, "count_holidays_between_leave"),
                "count_prefix_weekend": checkbox(data, "count_prefix_weekend"),
                "count_suffix_weekend": checkbox(data, "count_suffix_weekend"),
                "count_prefix_holiday": checkbox(data, "count_prefix_holiday"),
                "count_suffix_holiday": checkbox(data, "count_suffix_holiday"),
                "enabled": checkbox(data, "sandwich_enabled"),
                "updated_by": user,
                "created_by": user,
            },
        )

        LeaveProofRule.objects.update_or_create(
            organization=organization,
            leave_type=leave_type,
            defaults={
                "proof_required": checkbox(data, "proof_required"),
                "required_after_days": as_decimal(data.get("required_after_days")),
                "allowed_file_types": [item.strip().lower() for item in (data.get("allowed_file_types") or "pdf,jpg,jpeg,png").split(",") if item.strip()],
                "max_file_size_mb": as_int(data.get("max_file_size_mb"), 5),
                "requires_manual_review": checkbox(data, "requires_manual_review"),
                "reviewer_role": data.get("reviewer_role") or "HR",
                "enabled": checkbox(data, "proof_enabled"),
                "updated_by": user,
                "created_by": user,
            },
        )

        LeaveIntegrationRule.objects.update_or_create(
            organization=organization,
            leave_type=leave_type,
            defaults={
                "attendance_sync_enabled": checkbox(data, "attendance_sync_enabled"),
                "prevent_attendance_conflict": checkbox(data, "prevent_attendance_conflict"),
                "adjust_late_marks": checkbox(data, "adjust_late_marks"),
                "comp_off_generation_enabled": checkbox(data, "comp_off_generation_enabled"),
                "payroll_sync_enabled": checkbox(data, "payroll_sync_enabled"),
                "generate_lop": checkbox(data, "generate_lop"),
                "negative_balance_deduction": checkbox(data, "negative_balance_deduction"),
                "enable_encashment": checkbox(data, "enable_encashment"),
                "final_settlement_encashment": checkbox(data, "final_settlement_encashment"),
                "updated_by": user,
                "created_by": user,
            },
        )

        LeavePolicyWorkbenchService._save_workflow(organization=organization, leave_type=leave_type, user=user, data=data)
        LeaveAuditService.record(
            organization=organization,
            entity=leave_type,
            action="POLICY_CREATED" if created else "POLICY_UPDATED",
            user=user,
            new_value={"code": leave_type.code, "name": leave_type.name, "version": leave_type.policy_version},
            request=request,
        )
        return leave_type, created

    @classmethod
    @transaction.atomic
    def import_excel(cls, *, organization, user, uploaded_file, request=None):
        if not uploaded_file:
            raise ValueError("Select an Excel file to import.")

        filename = (uploaded_file.name or "").lower()
        if not filename.endswith(".xlsx"):
            raise ValueError("Upload a .xlsx Excel file.")

        from openpyxl import load_workbook

        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        worksheet = workbook.active
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = [cls._normalize_header(value) for value in (header_row or [])]
        if not headers or not any(headers):
            raise ValueError("Excel file must include a header row.")

        header_set = {header for header in headers if header}
        missing = cls.EXCEL_REQUIRED_COLUMNS - header_set
        if missing:
            raise ValueError(f"Excel file is missing required column(s): {', '.join(sorted(missing))}.")

        imported = 0
        errors = []
        for row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            raw_row = {
                headers[index]: values[index]
                for index in range(min(len(headers), len(values)))
                if headers[index]
            }
            if not any(value not in [None, ""] for value in raw_row.values()):
                continue

            try:
                row_data = cls._row_to_multivalue_dict(raw_row)
                cls.save_from_post(
                    organization=organization,
                    user=user,
                    data=row_data,
                    request=request,
                )
                imported += 1
            except Exception as exc:
                errors.append(f"Row {row_number}: {exc}")

        if errors:
            raise ValueError(" ".join(errors[:5]))
        if imported == 0:
            raise ValueError("No leave policies were found in the Excel file.")

        return imported

    @classmethod
    def sample_rows(cls):
        return [
            {
                "name": "Casual Leave",
                "code": "CL",
                "category_name": "Paid Leave",
                "category_code": "PAID",
                "status": "ACTIVE",
                "description": "General personal leave.",
                "max_balance": "12",
                "is_paid": "yes",
                "is_requestable": "yes",
                "accrual_frequency": "MONTHLY",
                "accrual_enabled": "yes",
                "accrual_rate": "1",
                "max_yearly_accrual": "12",
                "min_service_days": "0",
                "carry_forward_allowed": "no",
                "carry_forward_limit": "0",
                "proof_required": "no",
                "approval_route": "MANAGER_HR",
                "attendance_sync_enabled": "yes",
                "payroll_sync_enabled": "yes",
            },
            {
                "name": "Sick Leave",
                "code": "SL",
                "category_name": "Paid Leave",
                "category_code": "PAID",
                "status": "ACTIVE",
                "description": "Medical leave for illness.",
                "max_balance": "10",
                "is_paid": "yes",
                "is_requestable": "yes",
                "accrual_frequency": "MONTHLY",
                "accrual_enabled": "yes",
                "accrual_rate": "0.83",
                "max_yearly_accrual": "10",
                "min_service_days": "0",
                "carry_forward_allowed": "no",
                "proof_required": "yes",
                "required_after_days": "2",
                "approval_route": "MANAGER_HR",
                "attendance_sync_enabled": "yes",
                "payroll_sync_enabled": "yes",
            },
        ]

    @staticmethod
    def _normalize_header(value):
        return str(value or "").strip().lower().replace(" ", "_")

    @classmethod
    def _row_to_multivalue_dict(cls, raw_row):
        normalized = dict(cls.DEFAULT_IMPORT_VALUES)
        for key, value in raw_row.items():
            if value in [None, ""]:
                continue
            normalized[key] = cls._normalize_cell_value(key, value)

        values = {}
        for key, value in normalized.items():
            if key in cls.LIST_COLUMNS:
                values[key] = cls._split_list(value)
            else:
                values[key] = [str(value)]
        return MultiValueDict(values)

    @classmethod
    def _normalize_cell_value(cls, key, value):
        if key in cls.BOOLEAN_COLUMNS:
            return "on" if cls._truthy(value) else ""
        if hasattr(value, "date") and callable(value.date):
            return value.date().isoformat()
        if hasattr(value, "isoformat"):
            return value.isoformat()
        return str(value).strip()

    @staticmethod
    def _truthy(value):
        if isinstance(value, bool):
            return value
        return str(value).strip().lower() in {"1", "true", "yes", "y", "on"}

    @staticmethod
    def _split_list(value):
        if isinstance(value, (list, tuple)):
            return [str(item).strip() for item in value if str(item).strip()]
        return [item.strip() for item in str(value or "").split(",") if item.strip()]

    @staticmethod
    def _save_workflow(*, organization, leave_type, user, data):
        route = data.get("approval_route") or "MANAGER_HR"
        workflow, _ = LeaveWorkflow.objects.update_or_create(
            organization=organization,
            leave_type=leave_type,
            is_default=True,
            defaults={
                "name": f"{leave_type.code} Default Approval",
                "auto_approval_enabled": route == "AUTO",
                "escalation_enabled": checkbox(data, "workflow_escalation_enabled"),
                "updated_by": user,
                "created_by": user,
            },
        )
        workflow.steps.all().delete()

        routes = {
            "AUTO": [("AUTO_APPROVE", 1)],
            "MANAGER": [("REPORTING_MANAGER", 1)],
            "HR": [("HR", 1)],
            "MANAGER_HR": [("REPORTING_MANAGER", 1), ("HR", 2)],
            "DEPT_HEAD_HR": [("DEPARTMENT_HEAD", 1), ("HR", 2)],
        }
        for approver_type, order in routes.get(route, routes["MANAGER_HR"]):
            LeaveWorkflowStep.objects.create(
                organization=organization,
                workflow=workflow,
                order=order,
                approver_type=approver_type,
                approval_mode=data.get("approval_mode") or "ANY_ONE",
                sla_hours=as_int(data.get("approval_sla_hours"), 24),
                delegation_allowed=checkbox(data, "delegation_allowed"),
                created_by=user,
            )
        return workflow
